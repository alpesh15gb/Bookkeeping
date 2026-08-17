#!/usr/bin/env python3
"""Convert legacy bookkeeping data to reviewable CSV / Excel.

Run the SAME import code the ApexBooks app uses in production against a
throwaway database, then dump the result to CSV. What you review in the
spreadsheets is exactly what the app would have imported — there is no
second mapping to drift.

Usage
-----
    python -m tools.legacy_to_csv vyapar  backup.vyb  --out ./migration
    python -m tools.legacy_to_csv tally   export.xml --out ./migration
    python -m tools.legacy_to_csv vyapar  backup.vyb  --out ./migration --xlsx

Output
------
    out/contacts.csv            parties (customers/vendors) + opening balances
    out/products.csv            stock items with HSN / GST / prices
    out/invoices.csv            sales documents (headers)
    out/invoice_lines.csv       sales line items
    out/bills.csv               purchase documents (headers)
    out/bill_lines.csv          purchase line items
    out/payments.csv            receipts (incoming)
    out/bill_payments.csv       disbursements (outgoing)
    out/payment_allocations.csv receipt -> invoice allocation
    out/bill_payment_allocations.csv  disbursement -> bill allocation
    out/expenses.csv            expenses
    out/stock_ledger.csv        inventory movement ledger
    out/accounts.csv            chart of accounts with opening balances
    out/proforma_invoices.csv   estimates/quotes
    out/summary.json            machine-readable counts + money totals
    out/report.txt              human-readable reconciliation report
    out/bundle.xlsx             (with --xlsx) review-only Excel workbook

CSV is the canonical format. The .xlsx is for human review only: Excel
silently mangles 15-digit GSTINs and reformats dates, so never re-import
from the .xlsx.
"""
from __future__ import annotations

import argparse
import io
import json
import os
import shutil
import sqlite3
import sys
import tempfile
import uuid
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

BACKEND_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BACKEND_DIR))

# ── Dump definitions ─────────────────────────────────────────────────────
# Each entry is a SELECT whose column names become the CSV headers.
# Human-useful joins (contact/product names) keep the files reviewable.
DUMP_SELECTS = {
    "contacts": """
        SELECT name, contact_type, gstin, pan, email, phone, state_code,
               billing_address, opening_balance
        FROM contacts
        WHERE deleted_at IS NULL
        ORDER BY name
    """,
    "products": """
        SELECT name, sku, hsn_sac, product_type, uom,
               sales_price, purchase_price, gst_rate,
               opening_stock, current_stock, reorder_level
        FROM products
        WHERE deleted_at IS NULL
        ORDER BY name
    """,
    "invoices": """
        SELECT inv.invoice_number, c.name AS customer,
               inv.issue_date, inv.due_date, inv.pos_state_code, inv.status,
               inv.subtotal, inv.discount_total, inv.cgst_amount,
               inv.sgst_amount, inv.igst_amount, inv.cess_amount,
               inv.round_off, inv.shipping_charges, inv.total, inv.amount_paid
        FROM invoices inv
        LEFT JOIN contacts c ON c.id = inv.contact_id
        WHERE inv.deleted_at IS NULL
        ORDER BY inv.issue_date, inv.invoice_number
    """,
    "invoice_lines": """
        SELECT inv.invoice_number, p.name AS product, il.description,
               il.hsn_sac, il.quantity, il.rate, il.discount, il.subtotal,
               il.gst_rate, il.cgst_amount, il.sgst_amount, il.igst_amount,
               il.total
        FROM invoice_lines il
        JOIN invoices inv ON inv.id = il.invoice_id
        LEFT JOIN products p ON p.id = il.product_id
        WHERE inv.deleted_at IS NULL
        ORDER BY inv.invoice_number, il.id
    """,
    "bills": """
        SELECT b.bill_number, c.name AS vendor,
               b.issue_date, b.due_date, b.pos_state_code, b.status,
               b.subtotal, b.discount_total, b.cgst_amount, b.sgst_amount,
               b.igst_amount, b.cess_amount, b.round_off, b.shipping_charges,
               b.total, b.amount_paid, b.tds_amount
        FROM bills b
        LEFT JOIN contacts c ON c.id = b.contact_id
        WHERE b.deleted_at IS NULL
        ORDER BY b.issue_date, b.bill_number
    """,
    "bill_lines": """
        SELECT b.bill_number, p.name AS product, bl.description,
               bl.hsn_sac, bl.quantity, bl.rate, bl.discount, bl.subtotal,
               bl.gst_rate, bl.cgst_amount, bl.sgst_amount, bl.igst_amount,
               bl.total
        FROM bill_lines bl
        JOIN bills b ON b.id = bl.bill_id
        LEFT JOIN products p ON p.id = bl.product_id
        WHERE b.deleted_at IS NULL
        ORDER BY b.bill_number, bl.id
    """,
    "payments": """
        SELECT pay.payment_number, c.name AS customer,
               pay.payment_date, pay.payment_mode, pay.amount,
               pay.reference_number, pay.status
        FROM payments pay
        LEFT JOIN contacts c ON c.id = pay.contact_id
        WHERE pay.deleted_at IS NULL
        ORDER BY pay.payment_date, pay.payment_number
    """,
    "bill_payments": """
        SELECT bp.payment_number, c.name AS vendor,
               bp.payment_date, bp.payment_mode, bp.amount,
               bp.reference_number, bp.status
        FROM bill_payments bp
        LEFT JOIN contacts c ON c.id = bp.contact_id
        WHERE bp.deleted_at IS NULL
        ORDER BY bp.payment_date, bp.payment_number
    """,
    "payment_allocations": """
        SELECT pay.payment_number, inv.invoice_number, alloc.amount
        FROM payment_allocations alloc
        JOIN payments pay ON pay.id = alloc.payment_id
        LEFT JOIN invoices inv ON inv.id = alloc.invoice_id
        WHERE pay.deleted_at IS NULL
        ORDER BY pay.payment_number, inv.invoice_number
    """,
    "bill_payment_allocations": """
        SELECT bp.payment_number, b.bill_number, alloc.amount
        FROM bill_payment_allocations alloc
        JOIN bill_payments bp ON bp.id = alloc.payment_id
        LEFT JOIN bills b ON b.id = alloc.bill_id
        WHERE bp.deleted_at IS NULL
        ORDER BY bp.payment_number, b.bill_number
    """,
    "expenses": """
        SELECT e.expense_number, ec.name AS category, e.expense_date,
               e.vendor_name, e.description, e.amount, e.gst_rate,
               e.cgst_amount, e.sgst_amount, e.igst_amount, e.total, e.status
        FROM expenses e
        LEFT JOIN expense_categories ec ON ec.id = e.expense_category_id
        WHERE e.deleted_at IS NULL
        ORDER BY e.expense_date, e.expense_number
    """,
    "stock_ledger": """
        SELECT p.name AS product, DATE(sl.created_at) AS entry_date,
               sl.reference_type, sl.quantity, sl.balance_quantity, sl.rate
        FROM stock_ledger sl
        LEFT JOIN products p ON p.id = sl.product_id
        ORDER BY sl.created_at, p.name
    """,
    "accounts": """
        SELECT code, name AS account_name, account_type, opening_balance
        FROM accounts
        WHERE deleted_at IS NULL
        ORDER BY account_type, code
    """,
    "proforma_invoices": """
        SELECT pi.proforma_number, c.name AS customer, pi.issue_date,
               pi.due_date, pi.status, pi.total
        FROM proforma_invoices pi
        LEFT JOIN contacts c ON c.id = pi.contact_id
        WHERE pi.deleted_at IS NULL
        ORDER BY pi.issue_date, pi.proforma_number
    """,
}

TOTAL_QUERIES = {
    "invoices_total": "SELECT COALESCE(SUM(total), 0) FROM invoices WHERE deleted_at IS NULL",
    "invoices_gst": (
        "SELECT COALESCE(SUM(cgst_amount + sgst_amount + igst_amount), 0) "
        "FROM invoices WHERE deleted_at IS NULL"
    ),
    "bills_total": "SELECT COALESCE(SUM(total), 0) FROM bills WHERE deleted_at IS NULL",
    "bills_gst": (
        "SELECT COALESCE(SUM(cgst_amount + sgst_amount + igst_amount), 0) "
        "FROM bills WHERE deleted_at IS NULL"
    ),
    "payments_received": "SELECT COALESCE(SUM(amount), 0) FROM payments WHERE deleted_at IS NULL AND status = 'ACTIVE'",
    "payments_made": "SELECT COALESCE(SUM(amount), 0) FROM bill_payments WHERE deleted_at IS NULL AND status = 'ACTIVE'",
    "expenses_total": "SELECT COALESCE(SUM(total), 0) FROM expenses WHERE deleted_at IS NULL",
    # Net opening balance using the accounting convention: ASSET/EXPENSE are
    # debit-natured (positive), LIABILITY/EQUITY/REVENUE are credit-natured.
    "opening_balance_total": (
        "SELECT COALESCE(SUM(CASE WHEN account_type IN ('ASSET', 'EXPENSE') "
        "THEN opening_balance ELSE -opening_balance END), 0) "
        "FROM accounts WHERE deleted_at IS NULL"
    ),
}


def _boot_app(tmp_db: str):
    """Boot the real FastAPI app against a throwaway SQLite DB (test-style)."""
    from src.core.config import settings

    settings.DATABASE_URL = f"sqlite:///{tmp_db}"
    settings.SECRET_KEY = "migration-tool-secret-key"
    settings.JWT_SECRET_KEY = "migration-tool-jwt-secret-key-32chars"
    settings.APP_ENV = "development"
    settings.SEED_ON_STARTUP = False
    settings.RATE_LIMIT_ENABLED = False
    settings.REQUIRE_IDEMPOTENCY_KEY = False

    # Keep Celery from trying to reach a broker (same as the test suite).
    try:
        from src.core.celery import celery_app as core_celery
        core_celery.conf.update(task_always_eager=False, task_eager_propagates=False,
                                broker_connection_retry_on_startup=False)
    except Exception:
        pass
    try:
        from src.workers.tasks import celery_app as tasks_celery
        tasks_celery.conf.update(task_always_eager=False, task_eager_propagates=False,
                                 broker_connection_retry_on_startup=False)
    except Exception:
        pass

    import src.core.database as db_mod
    from sqlalchemy import create_engine, event
    from sqlalchemy.orm import sessionmaker
    from sqlalchemy.pool import StaticPool

    db_mod.engine = create_engine(
        settings.DATABASE_URL,
        poolclass=StaticPool,
        connect_args={"check_same_thread": False, "timeout": 30},
    )

    @event.listens_for(db_mod.engine, "connect")
    def _sqlite_pragmas(dbapi_conn, _record):
        cur = dbapi_conn.cursor()
        cur.execute("PRAGMA journal_mode=WAL")
        cur.execute("PRAGMA busy_timeout=30000")
        cur.close()

    db_mod.SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=db_mod.engine)

    from src.core.database import Base, get_db_session
    from src.main import app
    from fastapi.testclient import TestClient

    Base.metadata.create_all(bind=db_mod.engine)

    def _override_get_db_session():
        db = db_mod.SessionLocal()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db_session] = _override_get_db_session
    return TestClient(app)


def _register_tenant(client) -> dict:
    email = f"mig-{uuid.uuid4().hex[:10]}@company.com"
    r = client.post("/api/v1/auth/register", json={
        "email": email,
        "password": "SecurePassword123!",
        "full_name": "Migration Tool",
        "phone_number": "+919999999999",
        "company_legal_name": "Migration Co Pvt Ltd",
        "company_gstin": "27AAAAA1111A1Z1",
        "company_pan": "AAAAA1111A",
    })
    if r.status_code != 201:
        raise RuntimeError(f"tenant register failed: {r.status_code} {r.text[:300]}")
    login = client.post("/api/v1/auth/login", json={
        "email": email, "password": "SecurePassword123!",
    })
    if login.status_code != 200:
        raise RuntimeError(f"login failed: {login.status_code} {login.text[:300]}")

    from src.core.database import SessionLocal
    from src.infrastructure.database.models import User, TenantMembership
    db = SessionLocal()
    try:
        user = db.query(User).filter(User.email == email).first()
        tenant_id = db.query(TenantMembership).filter(
            TenantMembership.user_id == user.id).first().tenant_id
    finally:
        db.close()
    return {
        "Authorization": f"Bearer {login.json()['access_token']}",
        "X-Tenant-ID": str(tenant_id),
    }


def _write_csv(conn, out_dir: Path, filename: str, select_sql: str) -> int:
    import csv
    rows = conn.execute(select_sql).fetchall()
    path = out_dir / filename
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh)
        if rows:
            writer.writerow(rows[0].keys())
            for row in rows:
                writer.writerow(["" if v is None else v for v in row])
        else:
            # Empty table: still emit the header row from the cursor.
            names = [d[0] for d in conn.execute(select_sql).description]
            writer.writerow(names)
    return len(rows)


def _report_text(summary: dict, import_summary: dict, warnings: list[str]) -> str:
    lines = [
        "ApexBooks legacy conversion report",
        f"Generated: {datetime.now().isoformat(timespec='seconds')}",
        "",
        "Row counts",
        "----------",
    ]
    for name in ("contacts", "products", "invoices", "invoice_lines", "bills",
                 "bill_lines", "payments", "bill_payments", "payment_allocations",
                 "bill_payment_allocations", "expenses", "stock_ledger",
                 "accounts", "proforma_invoices"):
        lines.append(f"  {name:<28} {summary['row_counts'].get(name, 0)}")
    lines += [
        "",
        "Money totals",
        "------------",
        f"  Invoices total              {summary['totals'].get('invoices_total', 0)}",
        f"  Invoices GST (CGST+SGST+IGST){summary['totals'].get('invoices_gst', 0)}",
        f"  Bills total                 {summary['totals'].get('bills_total', 0)}",
        f"  Bills GST                   {summary['totals'].get('bills_gst', 0)}",
        f"  Payments received (active)  {summary['totals'].get('payments_received', 0)}",
        f"  Payments made (active)      {summary['totals'].get('payments_made', 0)}",
        f"  Expenses total              {summary['totals'].get('expenses_total', 0)}",
        f"  Opening balance net         {summary['totals'].get('opening_balance_total', 0)}",
        "",
        "Importer summary (from the app itself)",
        "-------------------------------------",
    ]
    for k, v in sorted(import_summary.items()):
        if k == "errors":
            continue
        lines.append(f"  {k:<32} {v}")
    if import_summary.get("errors"):
        lines += ["", "Import errors", "------------"] + [f"  - {e}" for e in import_summary["errors"]]
    if warnings:
        lines += ["", "Warnings", "--------"] + [f"  - {w}" for w in warnings]
    lines += [
        "",
        "Checks",
        "------",
        "  1. Open each CSV in Excel and spot-check a sample of rows.",
        "  2. Reconcile: invoices total and bills total above should match your",
        "     last GST return / closing balance in the legacy software.",
        "  3. Opening balance net should equal the legacy trial balance.",
        "  4. Keep the CSVs; they are the canonical migration artifact.",
    ]
    return "\n".join(lines) + "\n"


def _write_xlsx(out_dir: Path, csv_files: list[Path], summary: dict) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    import csv as _csv

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["ApexBooks migration — review only. CSV is canonical."])
    ws_summary.append([])
    for k, v in summary["totals"].items():
        ws_summary.append([k, v])
    ws_summary.append([])
    for k, v in summary["row_counts"].items():
        ws_summary.append([f"{k} rows", v])

    for path in csv_files:
        sheet_name = path.stem[:31]  # Excel sheet names <= 31 chars
        ws = wb.create_sheet(title=sheet_name)
        with open(path, newline="", encoding="utf-8-sig") as fh:
            reader = _csv.reader(fh)
            for ri, row in enumerate(reader):
                ws.append(row)
                if ri == 0:
                    for cell in ws[1]:
                        cell.font = Font(bold=True)
        # GSTIN / PAN / phone / document-number columns as text so Excel
        # never turns them into numbers or scientific notation.
        text_cols = {"gstin", "pan", "phone", "state_code", "hsn_sac", "uom",
                     "account_code", "invoice_number", "bill_number",
                     "payment_number", "expense_number", "proforma_number"}
        if ws.max_row > 1:
            headers = [c.value for c in ws[1]]
            for ci, h in enumerate(headers, start=1):
                if h in text_cols:
                    for cell in ws.iter_cols(min_col=ci, max_col=ci):
                        for c in cell[1:]:
                            c.number_format = "@"

    wb.save(out_dir / "bundle.xlsx")


def convert(source_format: str, source_file: Path, out_dir: Path, with_xlsx: bool) -> int:
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    warnings: list[str] = []

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".db")
    tmp_db = tmp.name
    tmp.close()
    try:
        client = _boot_app(tmp_db)
        headers = _register_tenant(client)

        data = source_file.read_bytes()
        if source_format == "vyapar":
            files = {"file": ("backup.vyb", data, "application/octet-stream")}
            resp = client.post("/api/v1/import/vyapar", files=files, headers=headers)
        else:
            files = {"file": ("export.xml", data, "application/xml")}
            resp = client.post("/api/v1/tally/import", files=files, headers=headers)

        import_summary = {}
        if resp.status_code == 200:
            try:
                import_summary = resp.json()
            except Exception:
                pass
        else:
            warnings.append(f"import endpoint returned {resp.status_code}: {resp.text[:500]}")

        conn = sqlite3.connect(tmp_db)
        conn.row_factory = sqlite3.Row
        try:
            row_counts = {}
            csv_files = []
            for name, select_sql in DUMP_SELECTS.items():
                try:
                    n = _write_csv(conn, out_dir, f"{name}.csv", select_sql)
                    row_counts[name] = n
                    csv_files.append(out_dir / f"{name}.csv")
                except Exception as exc:  # table may legitimately not exist
                    warnings.append(f"could not dump {name}: {exc}")
                    row_counts[name] = 0

            totals = {}
            for name, sql in TOTAL_QUERIES.items():
                try:
                    raw = conn.execute(sql).fetchone()[0] or 0
                    totals[name] = str(Decimal(str(raw)).quantize(Decimal("0.01")))
                except Exception as exc:
                    totals[name] = "0"
                    warnings.append(f"could not total {name}: {exc}")
        finally:
            conn.close()

        summary = {
            "source_format": source_format,
            "source_file": str(source_file),
            "import_status": resp.status_code,
            "import_summary": import_summary,
            "row_counts": row_counts,
            "totals": totals,
            "warnings": warnings,
        }
        (out_dir / "summary.json").write_text(
            json.dumps(summary, indent=2, default=str), encoding="utf-8")
        (out_dir / "report.txt").write_text(
            _report_text(summary, import_summary, warnings), encoding="utf-8")

        if with_xlsx:
            try:
                _write_xlsx(out_dir, csv_files, summary)
            except Exception as exc:
                warnings.append(f"xlsx bundle failed: {exc}")

        print(_report_text(summary, import_summary, warnings))
        return 0 if resp.status_code == 200 else 1
    finally:
        for suffix in ("", "-wal", "-shm"):
            try:
                os.remove(tmp_db + suffix)
            except OSError:
                pass


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(
        prog="legacy_to_csv",
        description=(
            "Convert Vyapar .vyb or Tally XML data to reviewable CSVs by running "
            "the app's real import code against a throwaway database."
        ),
    )
    sub = parser.add_subparsers(dest="source_format", required=True)

    p_vy = sub.add_parser("vyapar", help="convert a Vyapar .vyb backup")
    p_vy.add_argument("file", type=Path)
    p_vy.add_argument("--out", type=Path, required=True)
    p_vy.add_argument("--xlsx", action="store_true", help="also emit a review-only Excel bundle")

    p_tl = sub.add_parser("tally", help="convert a Tally XML export")
    p_tl.add_argument("file", type=Path)
    p_tl.add_argument("--out", type=Path, required=True)
    p_tl.add_argument("--xlsx", action="store_true", help="also emit a review-only Excel bundle")

    args = parser.parse_args(argv)
    if not Path(args.file).exists():
        print(f"error: file not found: {args.file}", file=sys.stderr)
        return 2
    return convert(args.source_format, Path(args.file), args.out, args.xlsx)


if __name__ == "__main__":
    sys.exit(main())
