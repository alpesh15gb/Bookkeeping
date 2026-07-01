"""
Reports API Router — Module 9: Reports & Analytics
All endpoints under /api/v1/reports/
"""
from fastapi import APIRouter, Depends, Query
from sqlalchemy.orm import Session
from typing import Optional
import uuid
from datetime import date

from src.core.database import get_db_session
from src.api.deps import enforce_permission
from src.infrastructure.database.models import Tenant
from io import BytesIO
from fastapi.responses import StreamingResponse
from src.domains.accounting.report_services import (
    BalanceSheetService,
    TrialBalanceService,
    GSTR1Service,
    GSTR3BService,
    AgingService,
    CashFlowService,
    SalesAnalyticsService,
    PurchaseAnalyticsService,
    OutstandingService,
    PartyStatementService,
)
from src.schemas.report_schemas import (
    BalanceSheetResponse,
    TrialBalanceResponse,
    GSTR1Response,
    GSTR3BResponse,
    AgingReportResponse,
    CashFlowResponse,
    SalesAnalyticsResponse,
    PurchaseAnalyticsResponse,
    OutstandingARResponse,
    OutstandingAPResponse,
    PartyStatementResponse,
    DayBookResponse,
    StockRegisterResponse,
    TDSReportResponse,
    TCSReportResponse
)
from src.schemas.gst_schemas import GSTR2Response

router = APIRouter(prefix="/reports", tags=["Reports & Analytics"])


# ---------------------------------------------------------------------------
# Balance Sheet
# ---------------------------------------------------------------------------

@router.get(
    "/balance-sheet",
    response_model=BalanceSheetResponse,
    summary="Balance Sheet",
    description=(
        "Returns the Balance Sheet (Assets = Liabilities + Equity) as of a given date. "
        "Current Year Net Profit/Loss is automatically computed from journal entries and "
        "injected into the Equity section."
    ),
)
def get_balance_sheet(
    as_of_date: date = Query(..., description="Report date, e.g. 2025-03-31"),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view")),
):
    return BalanceSheetService.get(db, tenant_id, as_of_date)


# ---------------------------------------------------------------------------
# Trial Balance
# ---------------------------------------------------------------------------

@router.get(
    "/trial-balance",
    response_model=TrialBalanceResponse,
    summary="Trial Balance",
    description=(
        "Returns the Trial Balance as of a given date. "
        "Shows all accounts (Asset, Liability, Equity, Revenue, Expense) with "
        "opening balance, period debits/credits, and closing balance."
    ),
)
def get_trial_balance(
    as_of_date: date = Query(..., description="Report date, e.g. 2025-03-31"),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view")),
):
    return TrialBalanceService.get(db, tenant_id, as_of_date)


# ---------------------------------------------------------------------------
# GST Reports
# ---------------------------------------------------------------------------

@router.get(
    "/gst/gstr1",
    response_model=GSTR1Response,
    summary="GSTR-1 Outward Supplies",
    description=(
        "Compiles GSTR-1 outward supply data for the given period. "
        "Splits invoices into B2B (registered), B2CL (inter-state large), "
        "B2CS (intra-state / small), and HSN-wise summary tables."
    ),
)
def get_gstr1(
    start_date: date = Query(..., description="Period start, e.g. 2025-04-01"),
    end_date: date = Query(..., description="Period end, e.g. 2025-06-30"),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view")),
):
    return GSTR1Service.get(db, tenant_id, start_date, end_date)


@router.get(
    "/gst/gstr2",
    response_model=GSTR2Response,
    summary="GSTR-2 Inward Supplies",
    description=(
        "Compiles GSTR-2 inward purchase data for the given period. "
        "Splits purchases into B2B (registered), B2BUR (reverse charge unregistered), "
        "and HSN-wise summary tables."
    ),
)
def get_gstr2(
    start_date: date = Query(..., description="Period start, e.g. 2025-04-01"),
    end_date: date = Query(..., description="Period end, e.g. 2025-06-30"),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view")),
):
    from src.api.v1.gst import get_gstr2_report
    return get_gstr2_report(start_date=start_date, end_date=end_date, db=db, tenant_id=tenant_id)


@router.get(
    "/gst/gstr3b",
    response_model=GSTR3BResponse,
    summary="GSTR-3B Monthly Summary",
    description=(
        "Compiles the GSTR-3B summary: outward taxable supplies, nil-rated supplies, "
        "and ITC available from purchase bills. Net tax payable = Output Tax − ITC."
    ),
)
def get_gstr3b(
    start_date: date = Query(..., description="Month start date, e.g. 2025-04-01"),
    end_date: date = Query(..., description="Month end date, e.g. 2025-04-30"),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view")),
):
    return GSTR3BService.get(db, tenant_id, start_date, end_date)



# ---------------------------------------------------------------------------
# Aging Reports
# ---------------------------------------------------------------------------

@router.get(
    "/aging/receivables",
    response_model=AgingReportResponse,
    summary="Accounts Receivable Aging",
    description=(
        "Groups outstanding customer invoices into aging buckets: "
        "0-30, 31-60, 61-90, and 91+ days overdue."
    ),
)
def get_ar_aging(
    as_of_date: date = Query(..., description="Report date for aging calculation"),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view")),
):
    return AgingService.get_receivables(db, tenant_id, as_of_date)


@router.get(
    "/aging/payables",
    response_model=AgingReportResponse,
    summary="Accounts Payable Aging",
    description=(
        "Groups outstanding vendor bills into aging buckets: "
        "0-30, 31-60, 61-90, and 91+ days overdue."
    ),
)
def get_ap_aging(
    as_of_date: date = Query(..., description="Report date for aging calculation"),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view")),
):
    return AgingService.get_payables(db, tenant_id, as_of_date)


# ---------------------------------------------------------------------------
# Cash Flow
# ---------------------------------------------------------------------------

@router.get(
    "/cash-flow",
    response_model=CashFlowResponse,
    summary="Cash Flow Statement",
    description=(
        "Indirect-method Cash Flow Statement. "
        "Operating = Net Profit ± changes in AR/AP. "
        "Investing = capital asset journal movements. "
        "Financing = equity/loan journal movements."
    ),
)
def get_cash_flow(
    start_date: date = Query(..., description="Period start"),
    end_date: date = Query(..., description="Period end"),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view")),
):
    return CashFlowService.get(db, tenant_id, start_date, end_date)


# ---------------------------------------------------------------------------
# Sales Analytics
# ---------------------------------------------------------------------------

@router.get(
    "/analytics/sales",
    response_model=SalesAnalyticsResponse,
    summary="Sales Analytics",
    description=(
        "Aggregate sales analytics for the period: total taxable sales, tax collected, "
        "invoice count, and top customers by invoice value."
    ),
)
def get_sales_analytics(
    start_date: date = Query(..., description="Period start"),
    end_date: date = Query(..., description="Period end"),
    top_n: int = Query(default=10, ge=1, le=50, description="Number of top customers to return"),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view")),
):
    return SalesAnalyticsService.get(db, tenant_id, start_date, end_date, top_n)


# ---------------------------------------------------------------------------
# Purchase Analytics
# ---------------------------------------------------------------------------

@router.get(
    "/analytics/purchases",
    response_model=PurchaseAnalyticsResponse,
    summary="Purchase Analytics",
    description=(
        "Aggregate purchase analytics for the period: total taxable purchases, tax paid, "
        "bill count, and top vendors by billed value."
    ),
)
def get_purchase_analytics(
    start_date: date = Query(..., description="Period start"),
    end_date: date = Query(..., description="Period end"),
    top_n: int = Query(default=10, ge=1, le=50, description="Number of top vendors to return"),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view")),
):
    return PurchaseAnalyticsService.get(db, tenant_id, start_date, end_date, top_n)


# ---------------------------------------------------------------------------
# Outstanding Documents (AR / AP snapshots)
# ---------------------------------------------------------------------------

@router.get(
    "/outstanding/receivables",
    response_model=OutstandingARResponse,
    summary="Outstanding Receivables",
    description="Lists all unpaid/partially-paid customer invoices with outstanding amounts as of a date.",
)
def get_outstanding_ar(
    as_of_date: date = Query(..., description="Snapshot date"),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view")),
):
    return OutstandingService.get_ar(db, tenant_id, as_of_date)


@router.get(
    "/outstanding/payables",
    response_model=OutstandingAPResponse,
    summary="Outstanding Payables",
    description="Lists all unpaid/partially-paid vendor bills with outstanding amounts as of a date.",
)
def get_outstanding_ap(
    as_of_date: date = Query(..., description="Snapshot date"),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view")),
):
    return OutstandingService.get_ap(db, tenant_id, as_of_date)


# ---------------------------------------------------------------------------
# Party Statement
# ---------------------------------------------------------------------------

@router.get(
    "/party-statement",
    response_model=PartyStatementResponse,
    summary="Party Statement JSON Data",
    description="Calculates and returns ledger, running balances, and summaries for a specific party.",
)
def get_party_statement(
    contact_id: uuid.UUID = Query(..., description="ID of the party/contact"),
    start_date: date = Query(..., description="Period start date"),
    end_date: date = Query(..., description="Period end date"),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view")),
):
    return PartyStatementService.get(db, tenant_id, contact_id, start_date, end_date)


@router.get(
    "/party-statement/pdf",
    summary="Party Statement A4 PDF",
    description="Generates and streams an A4 PDF document for the party statement.",
)
def get_party_statement_pdf(
    contact_id: uuid.UUID = Query(..., description="ID of the party/contact"),
    start_date: date = Query(..., description="Period start date"),
    end_date: date = Query(..., description="Period end date"),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view")),
):
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    from src.domains.printing.invoice_pdf import generate_party_statement_pdf
    from src.infrastructure.database.models import Tenant, TenantSetting

    statement = PartyStatementService.get(db, tenant_id, contact_id, start_date, end_date)
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    setting = db.query(TenantSetting).filter(TenantSetting.tenant_id == tenant_id).first()

    company_name = tenant.legal_name if tenant else "ApexBooks"
    company_gstin = tenant.gstin if tenant else None
    
    company_address = None
    company_phone = None
    if setting and setting.extra_settings:
        company_address = setting.extra_settings.get("company_address")
        company_phone = setting.extra_settings.get("company_phone")

    pdf_bytes = generate_party_statement_pdf(
        statement=statement,
        company_name=company_name,
        company_address=company_address,
        company_gstin=company_gstin,
        company_phone=company_phone,
    )

    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=PartyStatement_{statement.contact_name.replace(' ', '_')}.pdf"}
    )


@router.get(
    "/party-statement/excel",
    summary="Party Statement Excel Sheet",
    description="Generates and streams an Excel spreadsheet (.xlsx) for the party statement.",
)
def get_party_statement_excel(
    contact_id: uuid.UUID = Query(..., description="ID of the party/contact"),
    start_date: date = Query(..., description="Period start date"),
    end_date: date = Query(..., description="Period end date"),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view")),
):
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    from src.infrastructure.database.models import Tenant
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    statement = PartyStatementService.get(db, tenant_id, contact_id, start_date, end_date)
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    company_name = tenant.legal_name if tenant else "ApexBooks"

    # Excel Generation Logic
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Party Statement"
    ws.views.sheetView[0].showGridLines = True

    title_font = Font(name="Calibri", size=16, bold=True, color="0F1B3D")
    section_font = Font(name="Calibri", size=11, bold=True)
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    
    header_fill = PatternFill(start_color="0F1B3D", end_color="0F1B3D", fill_type="solid")
    accent_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    
    thin_side = Side(border_style="thin", color="D1D5DB")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    ws["A1"] = "Party Statement"
    ws["A1"].font = title_font
    
    ws["A3"] = f"Company Name: {company_name}"
    ws["A3"].font = bold_font
    
    ws["A4"] = f"Party Name: {statement.contact_name}"
    ws["A4"].font = bold_font
    ws["A5"] = f"Address: {statement.address or 'N/A'}"
    ws["A5"].font = normal_font
    ws["A6"] = f"GSTIN: {statement.gstin or 'N/A'}"
    ws["A6"].font = normal_font
    ws["A7"] = f"Mobile: {statement.phone or 'N/A'}"
    ws["A7"].font = normal_font

    start_str = statement.start_date.strftime("%d-%b-%Y")
    end_str = statement.end_date.strftime("%d-%b-%Y")
    ws["A9"] = f"Statement Period: {start_str} to {end_str}"
    ws["A9"].font = section_font

    headers = ["Date", "Particulars", "Voucher Type", "Voucher No.", "Debit (₹)", "Credit (₹)", "Balance (₹)"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=11, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="right" if col_idx in [5, 6, 7] else "left", vertical="center")

    current_row = 12
    for row in statement.ledger:
        date_str = row.date.strftime("%d-%b-%Y")
        ws.cell(row=current_row, column=1, value=date_str).font = normal_font
        ws.cell(row=current_row, column=2, value=row.particulars).font = normal_font
        ws.cell(row=current_row, column=3, value=row.voucher_type).font = normal_font
        ws.cell(row=current_row, column=4, value=row.voucher_no).font = normal_font
        
        deb_cell = ws.cell(row=current_row, column=5, value=float(row.debit) if row.debit is not None else "")
        deb_cell.font = normal_font
        deb_cell.number_format = "#,##0.00"
        
        cred_cell = ws.cell(row=current_row, column=6, value=float(row.credit) if row.credit is not None else "")
        cred_cell.font = normal_font
        cred_cell.number_format = "#,##0.00"
        
        bal_cell = ws.cell(row=current_row, column=7, value=row.balance)
        bal_cell.font = normal_font
        bal_cell.alignment = Alignment(horizontal="right")
        
        for c in range(1, 8):
            ws.cell(row=current_row, column=c).border = thin_border
            
        current_row += 1

    current_row += 2
    ws.cell(row=current_row, column=1, value="Summary").font = section_font
    current_row += 1
    
    sum_headers = ["Particulars", "Amount (₹)"]
    for col_idx, h in enumerate(sum_headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="right" if col_idx == 2 else "left", vertical="center")
    
    summary = statement.summary
    summary_rows = [
        ("Opening Balance", float(summary.opening_balance)),
    ]
    if statement.contact_type in ["CUSTOMER", "BOTH"] or summary.total_sales > 0 or summary.total_receipts > 0:
        summary_rows.append(("Total Sales", float(summary.total_sales)))
        summary_rows.append(("Total Receipts", float(summary.total_receipts)))
    if statement.contact_type in ["VENDOR", "BOTH"] or summary.total_purchases > 0 or summary.total_payments > 0:
        summary_rows.append(("Total Purchases", float(summary.total_purchases)))
        summary_rows.append(("Total Payments", float(summary.total_payments)))
    summary_rows.append(("Closing Outstanding", float(summary.closing_outstanding)))

    for label, val in summary_rows:
        current_row += 1
        is_closing = (label == "Closing Outstanding")
        lbl_cell = ws.cell(row=current_row, column=1, value=label)
        val_cell = ws.cell(row=current_row, column=2, value=val)
        
        lbl_cell.font = bold_font if is_closing else normal_font
        val_cell.font = bold_font if is_closing else normal_font
        val_cell.number_format = "#,##0.00"
        
        for c in [1, 2]:
            cell = ws.cell(row=current_row, column=c)
            cell.border = thin_border
            if is_closing:
                cell.fill = accent_fill

    # Adjust columns
    for col in ws.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            if cell.number_format and isinstance(cell.value, (int, float)):
                val_str = f"{cell.value:,.2f}"
            max_len = max(max_len, len(val_str))
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=PartyStatement_{statement.contact_name.replace(' ', '_')}.xlsx"}
    )


# --- Balance Sheet Exports ---
@router.get("/balance-sheet/excel")
def balance_sheet_excel(
    as_of_date: date = Query(..., description="Report date, e.g. 2025-03-31"),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view")),
):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    data = BalanceSheetService.get(db, tenant_id, as_of_date)
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    company_name = tenant.legal_name if tenant else "ApexBooks"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    ws.views.sheetView[0].showGridLines = True

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    header_fill = PatternFill(start_color="0F1B3D", end_color="0F1B3D", fill_type="solid")
    thin_side = Side(border_style="thin", color="D1D5DB")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    ws["A1"] = "Balance Sheet"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="0F1B3D")
    ws["A2"] = f"Company: {company_name} | As on: {as_of_date.strftime('%d-%b-%Y')}"
    ws["A2"].font = bold_font

    sections = [
        ("Assets", data.assets.items, data.assets.total),
        ("Liabilities", data.liabilities.items, data.liabilities.total),
        ("Equity", data.equity.items, data.equity.total)
    ]

    current_row = 4
    for sec_name, items, total in sections:
        cell = ws.cell(row=current_row, column=1, value=sec_name.upper())
        cell.font = header_font
        cell.fill = header_fill
        ws.cell(row=current_row, column=2, value="").fill = header_fill
        current_row += 1

        for item in items:
            ws.cell(row=current_row, column=1, value=f"{item.account_name} ({item.account_code})").font = normal_font
            val_cell = ws.cell(row=current_row, column=2, value=float(item.balance))
            val_cell.font = normal_font
            val_cell.number_format = "#,##0.00"
            current_row += 1

        tot_lbl = ws.cell(row=current_row, column=1, value=f"Total {sec_name}")
        tot_lbl.font = bold_font
        tot_val = ws.cell(row=current_row, column=2, value=float(total))
        tot_val.font = bold_font
        tot_val.number_format = "#,##0.00"
        current_row += 2

    # Column widths
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 20

    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=BalanceSheet_{as_of_date}.xlsx"}
    )


@router.get("/balance-sheet/pdf")
def balance_sheet_pdf(
    as_of_date: date = Query(..., description="Report date, e.g. 2025-03-31"),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view")),
):
    from src.domains.printing.invoice_pdf import generate_balance_sheet_pdf
    from fastapi.responses import StreamingResponse
    data = BalanceSheetService.get(db, tenant_id, as_of_date)
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    company_name = tenant.legal_name if tenant else "ApexBooks"

    pdf_bytes = generate_balance_sheet_pdf(
        data=data.model_dump(),
        company_name=company_name,
        cutoff=as_of_date.strftime("%d-%b-%Y")
    )
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=BalanceSheet_{as_of_date}.pdf"}
    )


# --- Profit & Loss Exports ---
@router.get("/profit-loss/excel")
def profit_loss_excel(
    start_date: date = Query(..., description="Start date"),
    end_date: date = Query(..., description="End date"),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view")),
):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from src.api.v1.accounting import get_profit_loss_report
    data = get_profit_loss_report(start_date, end_date, db, tenant_id)
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    company_name = tenant.legal_name if tenant else "ApexBooks"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Profit & Loss"
    ws.views.sheetView[0].showGridLines = True

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    header_fill = PatternFill(start_color="0F1B3D", end_color="0F1B3D", fill_type="solid")

    ws["A1"] = "Profit & Loss Statement"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="0F1B3D")
    ws["A2"] = f"Company: {company_name} | Period: {start_date.strftime('%d-%b-%Y')} to {end_date.strftime('%d-%b-%Y')}"
    ws["A2"].font = bold_font

    # Revenue
    ws.cell(row=4, column=1, value="REVENUE").font = header_font
    ws.cell(row=4, column=1).fill = header_fill
    ws.cell(row=4, column=2).fill = header_fill
    current_row = 5
    for item in data.revenue_lines:
        ws.cell(row=current_row, column=1, value=item.account_name).font = normal_font
        val = ws.cell(row=current_row, column=2, value=float(item.amount))
        val.font = normal_font
        val.number_format = "#,##0.00"
        current_row += 1
    
    ws.cell(row=current_row, column=1, value="Total Revenue").font = bold_font
    ws.cell(row=current_row, column=2, value=float(data.total_revenue)).font = bold_font
    ws.cell(row=current_row, column=2).number_format = "#,##0.00"
    current_row += 2

    # Expenses
    ws.cell(row=current_row, column=1, value="EXPENSES").font = header_font
    ws.cell(row=current_row, column=1).fill = header_fill
    ws.cell(row=current_row, column=2).fill = header_fill
    current_row += 1
    for item in data.expense_lines:
        ws.cell(row=current_row, column=1, value=item.account_name).font = normal_font
        val = ws.cell(row=current_row, column=2, value=float(item.amount))
        val.font = normal_font
        val.number_format = "#,##0.00"
        current_row += 1
    
    ws.cell(row=current_row, column=1, value="Total Expenses").font = bold_font
    ws.cell(row=current_row, column=2, value=float(data.total_expenses)).font = bold_font
    ws.cell(row=current_row, column=2).number_format = "#,##0.00"
    current_row += 2

    # Net Profit
    ws.cell(row=current_row, column=1, value="NET PROFIT / (LOSS)").font = bold_font
    ws.cell(row=current_row, column=2, value=float(data.net_profit)).font = bold_font
    ws.cell(row=current_row, column=2).number_format = "#,##0.00"

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 20

    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=ProfitLoss_{start_date}_to_{end_date}.xlsx"}
    )


@router.get("/profit-loss/pdf")
def profit_loss_pdf(
    start_date: date = Query(..., description="Start date"),
    end_date: date = Query(..., description="End date"),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view")),
):
    from src.domains.printing.invoice_pdf import generate_profit_loss_pdf
    from src.api.v1.accounting import get_profit_loss_report
    data = get_profit_loss_report(start_date, end_date, db, tenant_id)
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    company_name = tenant.legal_name if tenant else "ApexBooks"

    pdf_bytes = generate_profit_loss_pdf(
        data=data.model_dump(),
        company_name=company_name,
        start=start_date.strftime("%d-%b-%Y"),
        end=end_date.strftime("%d-%b-%Y")
    )
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=ProfitLoss_{start_date}_to_{end_date}.pdf"}
    )


# --- Trial Balance Exports ---
@router.get("/trial-balance/excel")
def trial_balance_excel(
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view")),
):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from src.api.v1.accounting import get_trial_balance
    from src.infrastructure.database.models import Tenant
    
    data = get_trial_balance(db=db, tenant_id=tenant_id)
    if not data or not hasattr(data, 'lines'):
        raise HTTPException(status_code=404, detail="Trial balance data not available")
    
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    company_name = tenant.legal_name if tenant else "ApexBooks"
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    company_name = tenant.legal_name if tenant else "ApexBooks"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trial Balance"
    ws.views.sheetView[0].showGridLines = True

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    header_fill = PatternFill(start_color="0F1B3D", end_color="0F1B3D", fill_type="solid")

    ws["A1"] = "Trial Balance"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="0F1B3D")
    ws["A2"] = f"Company: {company_name} | As on: {date.today().strftime('%d-%b-%Y')}"
    ws["A2"].font = bold_font

    headers = ["Account Code", "Account Name", "Opening Balance (₹)", "Total Debits (₹)", "Total Credits (₹)", "Closing Balance (₹)"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="right" if col_idx > 2 else "left")

    current_row = 5
    for line in data.lines:
        ws.cell(row=current_row, column=1, value=line.account_code).font = normal_font
        ws.cell(row=current_row, column=2, value=line.account_name).font = normal_font
        for col_idx, val in enumerate([line.opening_balance, line.total_debits, line.total_credits, line.closing_balance], start=3):
            cell = ws.cell(row=current_row, column=col_idx, value=float(val))
            cell.font = normal_font
            cell.number_format = "#,##0.00"
        current_row += 1

    ws.cell(row=current_row, column=2, value="Total").font = bold_font
    ws.cell(row=current_row, column=3, value=float(data.total_opening_debits)).font = bold_font
    ws.cell(row=current_row, column=3).number_format = "#,##0.00"
    ws.cell(row=current_row, column=4, value=float(data.total_debits)).font = bold_font
    ws.cell(row=current_row, column=4).number_format = "#,##0.00"
    ws.cell(row=current_row, column=5, value=float(data.total_credits)).font = bold_font
    ws.cell(row=current_row, column=5).number_format = "#,##0.00"
    ws.cell(row=current_row, column=6, value=float(data.total_closing_debits)).font = bold_font
    ws.cell(row=current_row, column=6).number_format = "#,##0.00"

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 30
    for c in ["C", "D", "E", "F"]:
        ws.column_dimensions[c].width = 20

    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=TrialBalance_{date.today()}.xlsx"}
    )


@router.get("/trial-balance/pdf")
def trial_balance_pdf(
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view")),
):
    from src.domains.printing.invoice_pdf import generate_trial_balance_pdf
    from src.api.v1.accounting import get_trial_balance
    data = get_trial_balance(db=db, tenant_id=tenant_id)
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    company_name = tenant.legal_name if tenant else "ApexBooks"

    pdf_bytes = generate_trial_balance_pdf(
        data=data.model_dump(),
        company_name=company_name
    )
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=TrialBalance_{date.today()}.pdf"}
    )


# --- Cash Flow Exports ---
@router.get("/cash-flow/excel")
def cash_flow_excel(
    start_date: date = Query(...),
    end_date: date = Query(...),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view")),
):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    data = CashFlowService.get(db, tenant_id, start_date, end_date)
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    company_name = tenant.legal_name if tenant else "ApexBooks"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Flow"
    ws.views.sheetView[0].showGridLines = True

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    header_fill = PatternFill(start_color="0F1B3D", end_color="0F1B3D", fill_type="solid")

    ws["A1"] = "Cash Flow Statement (Indirect Method)"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="0F1B3D")
    ws["A2"] = f"Company: {company_name} | Period: {start_date.strftime('%d-%b-%Y')} to {end_date.strftime('%d-%b-%Y')}"
    ws["A2"].font = bold_font

    sections = [
        ("Operating Activities", data.operating_activities.items, data.operating_activities.net),
        ("Investing Activities", data.investing_activities.items, data.investing_activities.net),
        ("Financing Activities", data.financing_activities.items, data.financing_activities.net)
    ]

    current_row = 4
    for sec_name, items, net in sections:
        cell = ws.cell(row=current_row, column=1, value=sec_name.upper())
        cell.font = header_font
        cell.fill = header_fill
        ws.cell(row=current_row, column=2, value="").fill = header_fill
        current_row += 1

        for item in items:
            ws.cell(row=current_row, column=1, value=item.label).font = normal_font
            val = ws.cell(row=current_row, column=2, value=float(item.amount))
            val.font = normal_font
            val.number_format = "#,##0.00"
            current_row += 1

        ws.cell(row=current_row, column=1, value=f"Net Cash from {sec_name}").font = bold_font
        val = ws.cell(row=current_row, column=2, value=float(net))
        val.font = bold_font
        val.number_format = "#,##0.00"
        current_row += 2

    # Reconciliation
    ws.cell(row=current_row, column=1, value="Net Change in Cash").font = bold_font
    ws.cell(row=current_row, column=2, value=float(data.net_change_in_cash)).font = bold_font
    ws.cell(row=current_row, column=2).number_format = "#,##0.00"
    current_row += 1

    ws.cell(row=current_row, column=1, value="Opening Cash Balance").font = normal_font
    ws.cell(row=current_row, column=2, value=float(data.opening_cash_balance)).font = normal_font
    ws.cell(row=current_row, column=2).number_format = "#,##0.00"
    current_row += 1

    ws.cell(row=current_row, column=1, value="Closing Cash Balance").font = bold_font
    ws.cell(row=current_row, column=2, value=float(data.closing_cash_balance)).font = bold_font
    ws.cell(row=current_row, column=2).number_format = "#,##0.00"

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 20

    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=CashFlow_{start_date}_to_{end_date}.xlsx"}
    )


@router.get("/cash-flow/pdf")
def cash_flow_pdf(
    start_date: date = Query(...),
    end_date: date = Query(...),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view")),
):
    from src.domains.printing.invoice_pdf import generate_cash_flow_pdf
    data = CashFlowService.get(db, tenant_id, start_date, end_date)
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    company_name = tenant.legal_name if tenant else "ApexBooks"

    pdf_bytes = generate_cash_flow_pdf(
        data=data.model_dump(),
        company_name=company_name,
        start=start_date.strftime("%d-%b-%Y"),
        end=end_date.strftime("%d-%b-%Y")
    )
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=CashFlow_{start_date}_to_{end_date}.pdf"}
    )


# --- Aging Report Exports ---
@router.get("/aging/{report_type}/excel")
def aging_excel(
    report_type: str,  # "receivables" or "payables"
    as_of_date: date = Query(...),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view")),
):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    if report_type == "receivables":
        data = AgingService.get_receivables(db, tenant_id, as_of_date)
    else:
        data = AgingService.get_payables(db, tenant_id, as_of_date)

    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    company_name = tenant.legal_name if tenant else "ApexBooks"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Aging Report"
    ws.views.sheetView[0].showGridLines = True

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    header_fill = PatternFill(start_color="0F1B3D", end_color="0F1B3D", fill_type="solid")

    ws["A1"] = f"{report_type.title()} Aging Report"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="0F1B3D")
    ws["A2"] = f"Company: {company_name} | As on: {as_of_date.strftime('%d-%b-%Y')}"
    ws["A2"].font = bold_font

    headers = ["Contact Name", "0-30 Days (₹)", "31-60 Days (₹)", "61-90 Days (₹)", "91+ Days (₹)", "Total Outstanding (₹)"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="right" if col_idx > 1 else "left")

    current_row = 5
    for line in data.lines:
        ws.cell(row=current_row, column=1, value=line.contact_name).font = normal_font
        for b_idx, bucket in enumerate(line.buckets, start=2):
            cell = ws.cell(row=current_row, column=b_idx, value=float(bucket.amount))
            cell.font = normal_font
            cell.number_format = "#,##0.00"
        
        tot = ws.cell(row=current_row, column=6, value=float(line.total_outstanding))
        tot.font = bold_font
        tot.number_format = "#,##0.00"
        current_row += 1

    # Totals Row
    ws.cell(row=current_row, column=1, value="Total").font = bold_font
    for b_idx, bucket in enumerate(data.bucket_totals, start=2):
        cell = ws.cell(row=current_row, column=b_idx, value=float(bucket.amount))
        cell.font = bold_font
        cell.number_format = "#,##0.00"
    
    tot_cell = ws.cell(row=current_row, column=6, value=float(data.total_outstanding))
    tot_cell.font = bold_font
    tot_cell.number_format = "#,##0.00"

    ws.column_dimensions["A"].width = 30
    for c in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[c].width = 18

    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Aging_{report_type}_{as_of_date}.xlsx"}
    )


@router.get("/aging/{report_type}/pdf")
def aging_pdf(
    report_type: str,
    as_of_date: date = Query(...),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view")),
):
    from src.domains.printing.invoice_pdf import generate_aging_pdf
    if report_type == "receivables":
        data = AgingService.get_receivables(db, tenant_id, as_of_date)
    else:
        data = AgingService.get_payables(db, tenant_id, as_of_date)
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    company_name = tenant.legal_name if tenant else "ApexBooks"

    pdf_bytes = generate_aging_pdf(
        data=data.model_dump(),
        company_name=company_name,
        as_of=as_of_date.strftime("%d-%b-%Y"),
        report_type=report_type
    )
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=Aging_{report_type}_{as_of_date}.pdf"}
    )


# --- Outstanding Document Exports ---
@router.get("/outstanding/{report_type}/excel")
def outstanding_excel(
    report_type: str,  # "receivables" or "payables"
    as_of_date: date = Query(...),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view")),
):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    if report_type == "receivables":
        data = OutstandingService.get_ar(db, tenant_id, as_of_date)
        items = data.invoices
        col_no_hdr = "Invoice Number"
    else:
        data = OutstandingService.get_ap(db, tenant_id, as_of_date)
        items = data.bills
        col_no_hdr = "Bill Number"

    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    company_name = tenant.legal_name if tenant else "ApexBooks"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Outstanding"
    ws.views.sheetView[0].showGridLines = True

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    header_fill = PatternFill(start_color="0F1B3D", end_color="0F1B3D", fill_type="solid")

    ws["A1"] = f"Outstanding {report_type.title()}"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="0F1B3D")
    ws["A2"] = f"Company: {company_name} | As on: {as_of_date.strftime('%d-%b-%Y')}"
    ws["A2"].font = bold_font

    headers = [col_no_hdr, "Contact Name", "Issue Date", "Due Date", "Total (₹)", "Paid (₹)", "Outstanding (₹)"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="right" if col_idx >= 5 else "left")

    current_row = 5
    for item in items:
        num = item.invoice_number if report_type == "receivables" else item.bill_number
        ws.cell(row=current_row, column=1, value=num).font = normal_font
        ws.cell(row=current_row, column=2, value=item.contact_name).font = normal_font
        ws.cell(row=current_row, column=3, value=item.issue_date.strftime("%d-%b-%Y")).font = normal_font
        ws.cell(row=current_row, column=4, value=item.due_date.strftime("%d-%b-%Y")).font = normal_font
        
        for c_idx, val in enumerate([item.total, item.amount_paid, item.outstanding], start=5):
            cell = ws.cell(row=current_row, column=c_idx, value=float(val))
            cell.font = normal_font
            cell.number_format = "#,##0.00"
        current_row += 1

    # Totals Row
    ws.cell(row=current_row, column=1, value="Total").font = bold_font
    tot_val = ws.cell(row=current_row, column=7, value=float(data.total_outstanding))
    tot_val.font = bold_font
    tot_val.number_format = "#,##0.00"

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 30
    for c in ["C", "D", "E", "F", "G"]:
        ws.column_dimensions[c].width = 16

    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Outstanding_{report_type}_{as_of_date}.xlsx"}
    )


@router.get("/outstanding/{report_type}/pdf")
def outstanding_pdf(
    report_type: str,
    as_of_date: date = Query(...),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view")),
):
    from src.domains.printing.invoice_pdf import generate_outstanding_pdf
    if report_type == "receivables":
        data = OutstandingService.get_ar(db, tenant_id, as_of_date)
    else:
        data = OutstandingService.get_ap(db, tenant_id, as_of_date)
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    company_name = tenant.legal_name if tenant else "ApexBooks"

    pdf_bytes = generate_outstanding_pdf(
        data=data.model_dump(),
        company_name=company_name,
        as_of=as_of_date.strftime("%d-%b-%Y"),
        report_type=report_type
    )
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=Outstanding_{report_type}_{as_of_date}.pdf"}
    )


# ---------------------------------------------------------------------------
# Cash Book
# ---------------------------------------------------------------------------

@router.get(
    "/cash-book",
    summary="Cash Book",
    description="Returns the Cash Book tracking cash inflows and outflows."
)
def get_cash_book(
    start_date: date = Query(..., description="Start date of the report"),
    end_date: date = Query(..., description="End date of the report"),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view")),
):
    from src.domains.accounting.report_services import CashBookService
    return CashBookService.get(db, tenant_id, start_date, end_date)

@router.get("/cash-book/excel")
def get_cash_book_excel(
    start_date: date = Query(...),
    end_date: date = Query(...),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view")),
):
    from fastapi.responses import Response
    from src.domains.accounting.report_services import CashBookService
    from src.domains.accounting.cash_book_export import generate_cash_book_excel
    
    data = CashBookService.get(db, tenant_id, start_date, end_date)
    excel_bytes = generate_cash_book_excel(data)
    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=CashBook_{start_date}_{end_date}.xlsx"}
    )

@router.get("/cash-book/pdf")
def get_cash_book_pdf(
    start_date: date = Query(...),
    end_date: date = Query(...),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view")),
):
    from fastapi.responses import Response
    from src.domains.accounting.report_services import CashBookService
    from src.domains.accounting.cash_book_export import generate_cash_book_pdf
    from src.infrastructure.database.models import Tenant
    
    tenant = db.query(Tenant).filter_by(id=tenant_id).first()
    company_name = tenant.legal_name if tenant else "ApexBooks"
    
    data = CashBookService.get(db, tenant_id, start_date, end_date)
    pdf_bytes = generate_cash_book_pdf(data, company_name)
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=CashBook_{start_date}_{end_date}.pdf"}
    )


# ── Day Book Exports & Endpoints ───────────────────────────────────────────

def generate_day_book_excel(data: dict, company_name: str) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Day Book"
    ws.views.sheetView[0].showGridLines = True

    # Title
    ws["A1"] = "Day Book"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="0F1B3D")
    ws["A2"] = f"Company: {company_name} | Period: {data['start_date']} to {data['end_date']}"
    ws["A2"].font = Font(name="Calibri", size=11, bold=True)

    # Headers
    headers = ["Date", "Voucher Type", "Voucher No.", "Particulars/Account", "Debit (₹)", "Credit (₹)", "Narration"]
    for col_num, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0F1B3D", end_color="0F1B3D", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    thin_side = Side(border_style="thin", color="D1D5DB")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    row_num = 5
    for entry in data["entries"]:
        first = True
        for line in entry["lines"]:
            if first:
                ws.cell(row=row_num, column=1, value=entry["entry_date"]).alignment = Alignment(horizontal="center")
                ws.cell(row=row_num, column=2, value=entry["source_type"]).alignment = Alignment(horizontal="center")
                ws.cell(row=row_num, column=3, value=entry["reference_number"]).alignment = Alignment(horizontal="center")
                first = False

            ws.cell(row=row_num, column=4, value=f"{line['account_name']} ({line['account_code']})")

            if line["direction"] == "DEBIT":
                ws.cell(row=row_num, column=5, value=float(line["amount"]))
            else:
                ws.cell(row=row_num, column=6, value=float(line["amount"]))

            ws.cell(row=row_num, column=7, value=line["narration"] or entry["description"])

            for col in range(1, 8):
                ws.cell(row=row_num, column=col).font = Font(name="Calibri", size=10)
                ws.cell(row=row_num, column=col).border = thin_border

            row_num += 1

    # Totals Row
    ws.cell(row=row_num, column=4, value="Total").font = Font(name="Calibri", size=11, bold=True)
    ws.cell(row=row_num, column=5, value=float(data["total_debit"])).font = Font(name="Calibri", size=11, bold=True)
    ws.cell(row=row_num, column=6, value=float(data["total_credit"])).font = Font(name="Calibri", size=11, bold=True)

    for col in range(1, 8):
        ws.cell(row=row_num, column=col).border = thin_border

    for col in range(1, 8):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def generate_day_book_pdf(data: dict, company_name: str) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    import io

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=15,
        leftMargin=15,
        topMargin=15,
        bottomMargin=15
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#0F1B3D')
    )
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#4B5563')
    )
    table_header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=colors.white
    )
    table_cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        leading=10
    )
    table_cell_bold = ParagraphStyle(
        'TableCellBold',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        leading=10
    )

    story = []

    story.append(Paragraph("Day Book", title_style))
    story.append(Spacer(1, 4))
    story.append(Paragraph(f"Company: {company_name} | Period: {data['start_date']} to {data['end_date']}", subtitle_style))
    story.append(Spacer(1, 15))

    headers = [
        Paragraph("Date", table_header_style),
        Paragraph("Type", table_header_style),
        Paragraph("Ref No.", table_header_style),
        Paragraph("Particulars/Account", table_header_style),
        Paragraph("Debit (₹)", table_header_style),
        Paragraph("Credit (₹)", table_header_style)
    ]

    table_data = [headers]

    for entry in data["entries"]:
        first = True
        for line in entry["lines"]:
            row = []
            if first:
                row.append(Paragraph(str(entry["entry_date"]), table_cell_style))
                row.append(Paragraph(str(entry["source_type"]), table_cell_style))
                row.append(Paragraph(str(entry["reference_number"] or ""), table_cell_style))
                first = False
            else:
                row.extend([Paragraph("", table_cell_style), Paragraph("", table_cell_style), Paragraph("", table_cell_style)])

            row.append(Paragraph(f"{line['account_name']} ({line['account_code']})", table_cell_style))

            if line["direction"] == "DEBIT":
                row.append(Paragraph(f"{line['amount']:.2f}", table_cell_style))
                row.append(Paragraph("", table_cell_style))
            else:
                row.append(Paragraph("", table_cell_style))
                row.append(Paragraph(f"{line['amount']:.2f}", table_cell_style))

            table_data.append(row)

    table_data.append([
        Paragraph("", table_cell_style),
        Paragraph("", table_cell_style),
        Paragraph("", table_cell_style),
        Paragraph("Total", table_cell_bold),
        Paragraph(f"{data['total_debit']:.2f}", table_cell_bold),
        Paragraph(f"{data['total_credit']:.2f}", table_cell_bold)
    ])

    col_widths = [60, 60, 65, 200, 90, 90]

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0F1B3D')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D1D5DB')),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#F3F4F6')),
    ]))

    story.append(t)
    doc.build(story)
    return buffer.getvalue()


@router.get("/day-book", response_model=DayBookResponse)
def get_day_book(
    start_date: date = Query(..., description="Start date of report"),
    end_date: date = Query(..., description="End date of report"),
    branch_id: Optional[uuid.UUID] = Query(None, description="Optional branch ID to filter by"),
    account_id: Optional[uuid.UUID] = Query(None, description="Optional ledger/account ID to filter by"),
    page: int = Query(default=1, ge=1),
    limit: int = Query(default=50, ge=1, le=200),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view"))
):
    from src.infrastructure.database.models import Branch, JournalEntry, JournalLine

    if branch_id:
        branch = db.query(Branch).filter(Branch.id == branch_id, Branch.tenant_id == tenant_id).first()
        if not branch:
            raise HTTPException(status_code=404, detail="Branch not found.")

    q = db.query(JournalEntry).filter(
        JournalEntry.tenant_id == tenant_id,
        JournalEntry.entry_date >= start_date,
        JournalEntry.entry_date <= end_date
    )

    if account_id:
        q = q.filter(JournalEntry.lines.any(JournalLine.account_id == account_id))

    total_count = q.count()

    offset = (page - 1) * limit
    journals = q.order_by(JournalEntry.entry_date.asc(), JournalEntry.created_at.asc()).offset(offset).limit(limit).all()

    all_journals = q.all()
    total_debit = Decimal("0.00")
    total_credit = Decimal("0.00")
    for j in all_journals:
        for l in j.lines:
            if l.direction == "DEBIT":
                total_debit += l.amount
            elif l.direction == "CREDIT":
                total_credit += l.amount

    entries = []
    for j in journals:
        lines = []
        for l in j.lines:
            lines.append({
                "account_id": str(l.account_id),
                "account_name": l.account.name if l.account else "Unknown",
                "account_code": l.account.code if l.account else "Unknown",
                "amount": l.amount,
                "direction": l.direction,
                "narration": l.narration
            })

        entries.append({
            "id": str(j.id),
            "entry_date": j.entry_date,
            "reference_number": j.reference_number or "",
            "description": j.description or "",
            "source_type": j.source_type,
            "source_id": str(j.source_id) if j.source_id else None,
            "lines": lines
        })

    return DayBookResponse(
        start_date=start_date,
        end_date=end_date,
        total_debit=total_debit,
        total_credit=total_credit,
        entries=entries,
        total_count=total_count
    )


@router.get("/day-book/excel")
def get_day_book_excel(
    start_date: date = Query(...),
    end_date: date = Query(...),
    branch_id: Optional[uuid.UUID] = Query(None),
    account_id: Optional[uuid.UUID] = Query(None),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view"))
):
    from fastapi.responses import Response
    from src.infrastructure.database.models import Tenant

    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    company_name = tenant.legal_name if tenant else "ApexBooks"

    day_book_raw = get_day_book(
        start_date=start_date,
        end_date=end_date,
        branch_id=branch_id,
        account_id=account_id,
        page=1,
        limit=10000,
        db=db,
        tenant_id=tenant_id
    )

    excel_bytes = generate_day_book_excel(day_book_raw.model_dump(), company_name)
    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=DayBook_{start_date}_{end_date}.xlsx"}
    )


@router.get("/day-book/pdf")
def get_day_book_pdf(
    start_date: date = Query(...),
    end_date: date = Query(...),
    branch_id: Optional[uuid.UUID] = Query(None),
    account_id: Optional[uuid.UUID] = Query(None),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view"))
):
    from fastapi.responses import Response
    from src.infrastructure.database.models import Tenant

    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    company_name = tenant.legal_name if tenant else "ApexBooks"

    day_book_raw = get_day_book(
        start_date=start_date,
        end_date=end_date,
        branch_id=branch_id,
        account_id=account_id,
        page=1,
        limit=10000,
        db=db,
        tenant_id=tenant_id
    )

    pdf_bytes = generate_day_book_pdf(day_book_raw.model_dump(), company_name)
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=DayBook_{start_date}_{end_date}.pdf"}
    )


@router.get("/stock-register", response_model=StockRegisterResponse)
def get_stock_register(
    start_date: Optional[date] = Query(None, description="Start date of report"),
    end_date: Optional[date] = Query(None, description="End date of report"),
    product_id: Optional[uuid.UUID] = Query(None, description="Optional product ID to filter by"),
    page: int = Query(default=1, ge=1),
    limit: int = Query(default=50, ge=1, le=200),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view"))
):
    from src.infrastructure.database.models import Product, StockLedger

    if not start_date:
        start_date = date(2026, 4, 1)
    if not end_date:
        end_date = date.today()

    pq = db.query(Product).filter(Product.tenant_id == tenant_id, Product.deleted_at == None)
    if product_id:
        pq = pq.filter(Product.id == product_id)

    total_count = pq.count()

    offset = (page - 1) * limit
    products = pq.order_by(Product.name.asc()).offset(offset).limit(limit).all()

    items = []
    for p in products:
        last_entry_before = db.query(StockLedger).filter(
            StockLedger.tenant_id == tenant_id,
            StockLedger.product_id == p.id,
            StockLedger.transaction_date < start_date
        ).order_by(StockLedger.transaction_date.desc(), StockLedger.created_at.desc()).first()

        opening_stock = last_entry_before.running_balance if last_entry_before else Decimal("0.00")

        entries = db.query(StockLedger).filter(
            StockLedger.tenant_id == tenant_id,
            StockLedger.product_id == p.id,
            StockLedger.transaction_date >= start_date,
            StockLedger.transaction_date <= end_date
        ).order_by(StockLedger.transaction_date.asc(), StockLedger.created_at.asc()).all()

        inward_qty = Decimal("0.00")
        outward_qty = Decimal("0.00")
        adjustment_qty = Decimal("0.00")

        for e in entries:
            if e.transaction_type == "ADJUSTMENT":
                adjustment_qty += e.quantity_change
            else:
                if e.quantity_change > 0:
                    inward_qty += e.quantity_change
                else:
                    outward_qty += abs(e.quantity_change)

        closing_stock = opening_stock + inward_qty - outward_qty + adjustment_qty

        items.append({
            "product_id": p.id,
            "product_name": p.name,
            "sku": p.sku,
            "uom": p.uom,
            "opening_stock": opening_stock,
            "inward_qty": inward_qty,
            "outward_qty": outward_qty,
            "adjustment_qty": adjustment_qty,
            "closing_stock": closing_stock
        })

    return StockRegisterResponse(
        start_date=start_date,
        end_date=end_date,
        items=items,
        total_count=total_count
    )


# ── TDS/TCS Reports Endpoints ───────────────────────────────────────────────

def generate_tcs_excel(data: dict, company_name: str) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TCS Report"
    ws.views.sheetView[0].showGridLines = True

    ws["A1"] = "TCS Report"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="0F1B3D")
    ws["A2"] = f"Company: {company_name} | Period: {data['start_date']} to {data['end_date']} | Type: {data['report_type'].upper()}"
    ws["A2"].font = Font(name="Calibri", size=11, bold=True)

    thin_side = Side(border_style="thin", color="D1D5DB")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    if data["report_type"] == "detailed":
        headers = ["Date", "Reference No.", "Party Name", "Taxable Amount (₹)", "TCS Rate (%)", "TCS Amount (₹)"]
        for col_num, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num, value=h)
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0F1B3D", end_color="0F1B3D", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        row_num = 5
        for item in data["detailed_items"]:
            ws.cell(row=row_num, column=1, value=str(item["date"])).alignment = Alignment(horizontal="center")
            ws.cell(row=row_num, column=2, value=item["number"]).alignment = Alignment(horizontal="center")
            ws.cell(row=row_num, column=3, value=item["contact_name"])
            ws.cell(row=row_num, column=4, value=float(item["taxable_amount"]))
            ws.cell(row=row_num, column=5, value=float(item["tcs_rate"]))
            ws.cell(row=row_num, column=6, value=float(item["tcs_amount"]))

            for col in range(1, 7):
                ws.cell(row=row_num, column=col).border = thin_border
                ws.cell(row=row_num, column=col).font = Font(name="Calibri", size=10)
            row_num += 1

        ws.cell(row=row_num, column=3, value="Total").font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=row_num, column=6, value=float(data["total_tcs_amount"])).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=row_num, column=3).border = thin_border
        ws.cell(row=row_num, column=6).border = thin_border

    else:
        headers = ["Party Name", "Total Taxable Amount (₹)", "Total TCS Amount (₹)"]
        for col_num, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num, value=h)
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0F1B3D", end_color="0F1B3D", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        row_num = 5
        for item in data["summary_items"]:
            ws.cell(row=row_num, column=1, value=item["contact_name"])
            ws.cell(row=row_num, column=2, value=float(item["total_taxable_amount"]))
            ws.cell(row=row_num, column=3, value=float(item["total_tcs_amount"]))

            for col in range(1, 4):
                ws.cell(row=row_num, column=col).border = thin_border
                ws.cell(row=row_num, column=col).font = Font(name="Calibri", size=10)
            row_num += 1

        ws.cell(row=row_num, column=1, value="Total").font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=row_num, column=3, value=float(data["total_tcs_amount"])).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=row_num, column=1).border = thin_border
        ws.cell(row=row_num, column=3).border = thin_border

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def generate_tds_pdf(data: dict, company_name: str) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    import io

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=15, leftMargin=15, topMargin=15, bottomMargin=15)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('DocTitle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=18, textColor=colors.HexColor('#0F1B3D'))
    subtitle_style = ParagraphStyle('DocSubtitle', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=10, textColor=colors.HexColor('#4B5563'))
    th_style = ParagraphStyle('TH', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=9, textColor=colors.white)
    tc_style = ParagraphStyle('TC', parent=styles['Normal'], fontName='Helvetica', fontSize=8)
    tc_bold = ParagraphStyle('TCB', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8)

    story = [
        Paragraph("TDS Report", title_style),
        Spacer(1, 4),
        Paragraph(f"Company: {company_name} | Period: {data['start_date']} to {data['end_date']} | Type: {data['report_type'].upper()}", subtitle_style),
        Spacer(1, 15)
    ]

    if data["report_type"] == "detailed":
        headers = [Paragraph("Date", th_style), Paragraph("Type", th_style), Paragraph("Ref No.", th_style), Paragraph("Party Name", th_style), Paragraph("Taxable (₹)", th_style), Paragraph("TDS Amount (₹)", th_style)]
        table_data = [headers]
        for item in data["detailed_items"]:
            table_data.append([
                Paragraph(str(item["date"]), tc_style),
                Paragraph(item["type"], tc_style),
                Paragraph(item["number"], tc_style),
                Paragraph(item["contact_name"], tc_style),
                Paragraph(f"{item['taxable_amount']:.2f}", tc_style),
                Paragraph(f"{item['tds_amount']:.2f}", tc_style)
            ])
        table_data.append([Paragraph("", tc_style), Paragraph("", tc_style), Paragraph("", tc_style), Paragraph("Total", tc_bold), Paragraph("", tc_bold), Paragraph(f"{data['total_tds_amount']:.2f}", tc_bold)])
        col_widths = [65, 55, 65, 180, 100, 100]
    else:
        headers = [Paragraph("Party Name", th_style), Paragraph("Total Taxable Amount (₹)", th_style), Paragraph("Total TDS Amount (₹)", th_style)]
        table_data = [headers]
        for item in data["summary_items"]:
            table_data.append([
                Paragraph(item["contact_name"], tc_style),
                Paragraph(f"{item['total_taxable_amount']:.2f}", tc_style),
                Paragraph(f"{item['total_tds_amount']:.2f}", tc_style)
            ])
        table_data.append([Paragraph("Total", tc_bold), Paragraph("", tc_bold), Paragraph(f"{data['total_tds_amount']:.2f}", tc_bold)])
        col_widths = [265, 150, 150]

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0F1B3D')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D1D5DB')),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#F3F4F6')),
    ]))
    story.append(t)
    doc.build(story)
    return buffer.getvalue()


def generate_tcs_pdf(data: dict, company_name: str) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    import io

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=15, leftMargin=15, topMargin=15, bottomMargin=15)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('DocTitle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=18, textColor=colors.HexColor('#0F1B3D'))
    subtitle_style = ParagraphStyle('DocSubtitle', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=10, textColor=colors.HexColor('#4B5563'))
    th_style = ParagraphStyle('TH', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=9, textColor=colors.white)
    tc_style = ParagraphStyle('TC', parent=styles['Normal'], fontName='Helvetica', fontSize=8)
    tc_bold = ParagraphStyle('TCB', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8)

    story = [
        Paragraph("TCS Report", title_style),
        Spacer(1, 4),
        Paragraph(f"Company: {company_name} | Period: {data['start_date']} to {data['end_date']} | Type: {data['report_type'].upper()}", subtitle_style),
        Spacer(1, 15)
    ]

    if data["report_type"] == "detailed":
        headers = [Paragraph("Date", th_style), Paragraph("Ref No.", th_style), Paragraph("Party Name", th_style), Paragraph("Taxable (₹)", th_style), Paragraph("TCS Amount (₹)", th_style)]
        table_data = [headers]
        for item in data["detailed_items"]:
            table_data.append([
                Paragraph(str(item["date"]), tc_style),
                Paragraph(item["number"], tc_style),
                Paragraph(item["contact_name"], tc_style),
                Paragraph(f"{item['taxable_amount']:.2f}", tc_style),
                Paragraph(f"{item['tcs_amount']:.2f}", tc_style)
            ])
        table_data.append([Paragraph("", tc_style), Paragraph("", tc_style), Paragraph("Total", tc_bold), Paragraph("", tc_bold), Paragraph(f"{data['total_tcs_amount']:.2f}", tc_bold)])
        col_widths = [80, 80, 205, 100, 100]
    else:
        headers = [Paragraph("Party Name", th_style), Paragraph("Total Taxable Amount (₹)", th_style), Paragraph("Total TCS Amount (₹)", th_style)]
        table_data = [headers]
        for item in data["summary_items"]:
            table_data.append([
                Paragraph(item["contact_name"], tc_style),
                Paragraph(f"{item['total_taxable_amount']:.2f}", tc_style),
                Paragraph(f"{item['total_tcs_amount']:.2f}", tc_style)
            ])
        table_data.append([Paragraph("Total", tc_bold), Paragraph("", tc_bold), Paragraph(f"{data['total_tcs_amount']:.2f}", tc_bold)])
        col_widths = [265, 150, 150]

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0F1B3D')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D1D5DB')),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#F3F4F6')),
    ]))
    story.append(t)
    doc.build(story)
    return buffer.getvalue()


@router.get("/tds", response_model=TDSReportResponse)
def get_tds_report(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    contact_id: Optional[uuid.UUID] = Query(None),
    report_type: str = Query("summary", pattern="^(summary|detailed)$"),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view"))
):
    from src.infrastructure.database.models import Invoice, Bill

    if not start_date:
        start_date = date(2026, 4, 1)
    if not end_date:
        end_date = date.today()

    bq = db.query(Bill).filter(Bill.tenant_id == tenant_id, Bill.deleted_at == None, Bill.tds_amount > 0)
    if contact_id:
        bq = bq.filter(Bill.contact_id == contact_id)
    bq = bq.filter(Bill.issue_date >= start_date, Bill.issue_date <= end_date)
    bills = bq.all()

    iq = db.query(Invoice).filter(Invoice.tenant_id == tenant_id, Invoice.deleted_at == None, Invoice.tds_amount > 0)
    if contact_id:
        iq = iq.filter(Invoice.contact_id == contact_id)
    iq = iq.filter(Invoice.issue_date >= start_date, Invoice.issue_date <= end_date)
    invoices = iq.all()

    detailed_items = []
    for b in bills:
        detailed_items.append({
            "id": b.id,
            "type": "BILL",
            "number": b.bill_number,
            "date": b.issue_date,
            "contact_name": b.contact.name if b.contact else "Unknown Vendor",
            "taxable_amount": b.subtotal,
            "tds_rate": b.tds_rate,
            "tds_amount": b.tds_amount
        })

    for inv in invoices:
        detailed_items.append({
            "id": inv.id,
            "type": "INVOICE",
            "number": inv.invoice_number,
            "date": inv.issue_date,
            "contact_name": inv.contact.name if inv.contact else "Unknown Customer",
            "taxable_amount": inv.subtotal,
            "tds_rate": inv.tds_rate,
            "tds_amount": inv.tds_amount
        })

    detailed_items.sort(key=lambda x: (x["date"], x["number"]))
    total_tds = sum(x["tds_amount"] for x in detailed_items)

    if report_type == "detailed":
        return TDSReportResponse(
            start_date=start_date,
            end_date=end_date,
            report_type=report_type,
            total_tds_amount=total_tds,
            detailed_items=detailed_items
        )

    summary_map = {}
    for x in detailed_items:
        c_name = x["contact_name"]
        c_id = uuid.UUID("00000000-0000-0000-0000-000000000000")
        matched_doc = db.query(Bill).filter(Bill.id == x["id"]).first() or db.query(Invoice).filter(Invoice.id == x["id"]).first()
        if matched_doc and matched_doc.contact_id:
            c_id = matched_doc.contact_id

        if c_name not in summary_map:
            summary_map[c_name] = {
                "contact_id": c_id,
                "contact_name": c_name,
                "total_taxable_amount": Decimal("0.00"),
                "total_tds_amount": Decimal("0.00")
            }
        summary_map[c_name]["total_taxable_amount"] += x["taxable_amount"]
        summary_map[c_name]["total_tds_amount"] += x["tds_amount"]

    summary_items = list(summary_map.values())
    summary_items.sort(key=lambda x: x["contact_name"])

    return TDSReportResponse(
        start_date=start_date,
        end_date=end_date,
        report_type=report_type,
        total_tds_amount=total_tds,
        summary_items=summary_items
    )


@router.get("/tcs", response_model=TCSReportResponse)
def get_tcs_report(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    contact_id: Optional[uuid.UUID] = Query(None),
    report_type: str = Query("summary", pattern="^(summary|detailed)$"),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view"))
):
    from src.infrastructure.database.models import Invoice

    if not start_date:
        start_date = date(2026, 4, 1)
    if not end_date:
        end_date = date.today()

    iq = db.query(Invoice).filter(Invoice.tenant_id == tenant_id, Invoice.deleted_at == None, Invoice.tcs_amount > 0)
    if contact_id:
        iq = iq.filter(Invoice.contact_id == contact_id)
    iq = iq.filter(Invoice.issue_date >= start_date, Invoice.issue_date <= end_date)
    invoices = iq.all()

    detailed_items = []
    for inv in invoices:
        detailed_items.append({
            "id": inv.id,
            "number": inv.invoice_number,
            "date": inv.issue_date,
            "contact_name": inv.contact.name if inv.contact else "Unknown Customer",
            "taxable_amount": inv.subtotal,
            "tcs_rate": inv.tcs_rate,
            "tcs_amount": inv.tcs_amount
        })

    detailed_items.sort(key=lambda x: (x["date"], x["number"]))
    total_tcs = sum(x["tcs_amount"] for x in detailed_items)

    if report_type == "detailed":
        return TCSReportResponse(
            start_date=start_date,
            end_date=end_date,
            report_type=report_type,
            total_tcs_amount=total_tcs,
            detailed_items=detailed_items
        )

    summary_map = {}
    for x in detailed_items:
        c_name = x["contact_name"]
        c_id = uuid.UUID("00000000-0000-0000-0000-000000000000")
        matched_doc = db.query(Invoice).filter(Invoice.id == x["id"]).first()
        if matched_doc and matched_doc.contact_id:
            c_id = matched_doc.contact_id

        if c_name not in summary_map:
            summary_map[c_name] = {
                "contact_id": c_id,
                "contact_name": c_name,
                "total_taxable_amount": Decimal("0.00"),
                "total_tcs_amount": Decimal("0.00")
            }
        summary_map[c_name]["total_taxable_amount"] += x["taxable_amount"]
        summary_map[c_name]["total_tcs_amount"] += x["tcs_amount"]

    summary_items = list(summary_map.values())
    summary_items.sort(key=lambda x: x["contact_name"])

    return TCSReportResponse(
        start_date=start_date,
        end_date=end_date,
        report_type=report_type,
        total_tcs_amount=total_tcs,
        summary_items=summary_items
    )


@router.get("/tds/excel")
def get_tds_report_excel(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    contact_id: Optional[uuid.UUID] = Query(None),
    report_type: str = Query("summary"),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view"))
):
    from fastapi.responses import Response
    from src.infrastructure.database.models import Tenant

    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    company_name = tenant.legal_name if tenant else "ApexBooks"

    data = get_tds_report(start_date=start_date, end_date=end_date, contact_id=contact_id, report_type=report_type, db=db, tenant_id=tenant_id)
    excel_bytes = generate_tds_excel(data.model_dump(), company_name)
    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=TDS_Report_{start_date}_{end_date}.xlsx"}
    )


@router.get("/tcs/excel")
def get_tcs_report_excel(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    contact_id: Optional[uuid.UUID] = Query(None),
    report_type: str = Query("summary"),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view"))
):
    from fastapi.responses import Response
    from src.infrastructure.database.models import Tenant

    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    company_name = tenant.legal_name if tenant else "ApexBooks"

    data = get_tcs_report(start_date=start_date, end_date=end_date, contact_id=contact_id, report_type=report_type, db=db, tenant_id=tenant_id)
    excel_bytes = generate_tcs_excel(data.model_dump(), company_name)
    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=TCS_Report_{start_date}_{end_date}.xlsx"}
    )


@router.get("/tds/pdf")
def get_tds_report_pdf(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    contact_id: Optional[uuid.UUID] = Query(None),
    report_type: str = Query("summary"),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view"))
):
    from fastapi.responses import Response
    from src.infrastructure.database.models import Tenant

    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    company_name = tenant.legal_name if tenant else "ApexBooks"

    data = get_tds_report(start_date=start_date, end_date=end_date, contact_id=contact_id, report_type=report_type, db=db, tenant_id=tenant_id)
    pdf_bytes = generate_tds_pdf(data.model_dump(), company_name)
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=TDS_Report_{start_date}_{end_date}.pdf"}
    )


@router.get("/tcs/pdf")
def get_tcs_report_pdf(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    contact_id: Optional[uuid.UUID] = Query(None),
    report_type: str = Query("summary"),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view"))
):
    from fastapi.responses import Response
    from src.infrastructure.database.models import Tenant

    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    company_name = tenant.legal_name if tenant else "ApexBooks"

    data = get_tcs_report(start_date=start_date, end_date=end_date, contact_id=contact_id, report_type=report_type, db=db, tenant_id=tenant_id)
    pdf_bytes = generate_tcs_pdf(data.model_dump(), company_name)
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=TCS_Report_{start_date}_{end_date}.pdf"}
    )
