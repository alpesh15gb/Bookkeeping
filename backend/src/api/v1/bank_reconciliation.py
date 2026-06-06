"""
Bank Reconciliation Module
==========================
Handles bank statement import (CSV/Excel), transaction parsing,
auto-matching with deduplication, bulk reconciliation, and reporting.
"""
import io
import csv
import uuid
import re
from decimal import Decimal, InvalidOperation
from datetime import date, datetime, timedelta
from typing import List, Optional, Dict, Tuple
from difflib import SequenceMatcher

from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Query
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func, and_, or_

from src.core.database import get_db_session
from src.infrastructure.database.models import (
    BankStatement, BankTransaction, BankReconciliation, BankingProfile,
    Payment, BillPayment, Invoice, Bill, Contact
)
from src.schemas.bill_schemas import (
    BankStatementCreate, BankStatementResponse, BankStatementListResponse,
    BankTransactionCreate, BankTransactionResponse,
    BankReconciliationCreate, BankReconciliationResponse, BankReconciliationListResponse
)
from src.api.deps import enforce_permission
from pydantic import BaseModel, Field

router = APIRouter(prefix="/bank-reconciliation", tags=["Bank Reconciliation"])


# ---------------------------------------------------------------------------
# NEW SCHEMAS
# ---------------------------------------------------------------------------

class BulkReconcileItem(BaseModel):
    transaction_id: uuid.UUID
    payment_id: Optional[uuid.UUID] = None
    bill_payment_id: Optional[uuid.UUID] = None
    amount: Decimal


class BulkReconcileRequest(BaseModel):
    items: List[BulkReconcileItem]


class MatchSuggestion(BaseModel):
    transaction_id: str
    transaction_date: str
    transaction_amount: str
    transaction_description: Optional[str] = None
    suggested_matches: List[dict] = []


class StatementStats(BaseModel):
    total_transactions: int = 0
    reconciled: int = 0
    pending: int = 0
    total_credits: Decimal = Decimal("0")
    total_debits: Decimal = Decimal("0")
    reconciliation_pct: float = 0.0


class PendingInvoice(BaseModel):
    id: str
    invoice_number: str
    contact_name: Optional[str] = None
    issue_date: str
    due_date: str
    total: Decimal
    amount_paid: Decimal
    outstanding: Decimal
    days_overdue: int = 0


class PendingBill(BaseModel):
    id: str
    bill_number: str
    contact_name: Optional[str] = None
    issue_date: str
    due_date: str
    total: Decimal
    amount_paid: Decimal
    outstanding: Decimal
    days_overdue: int = 0


# ---------------------------------------------------------------------------
# BANK STATEMENT PARSING
# ---------------------------------------------------------------------------

# Common Indian bank CSV column mappings
BANK_FORMATS = {
    "SBI": {
        "date_cols": ["Txn Date", "Transaction Date", "Date"],
        "desc_cols": ["Description", "Narration", "Remarks"],
        "ref_cols": ["Cheque No.", "Cheque No", "Reference No", "Ref No"],
        "debit_cols": ["Debit", "Debit Amount", "Withdrawal", "DR"],
        "credit_cols": ["Credit", "Credit Amount", "Deposit", "CR"],
        "balance_cols": ["Balance", "Closing Balance", "Running Balance"],
        "date_formats": ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"],
    },
    "HDFC": {
        "date_cols": ["Date", "Transaction Date", "Txn Date"],
        "desc_cols": ["Narration", "Description", "Remarks"],
        "ref_cols": ["Chq/Ref No", "Reference", "Cheque No"],
        "debit_cols": ["Debit Amount", "Debit", "Withdrawal", "DR"],
        "credit_cols": ["Credit Amount", "Credit", "Deposit", "CR"],
        "balance_cols": ["Balance", "Closing Balance"],
        "date_formats": ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"],
    },
    "ICICI": {
        "date_cols": ["Date", "Value Date", "Transaction Date"],
        "desc_cols": ["Description", "Narration", "Remarks"],
        "ref_cols": ["Reference", "Ref No", "Cheque No"],
        "debit_cols": ["Debit Amount", "Debit", "Withdrawal", "DR"],
        "credit_cols": ["Credit Amount", "Credit", "Deposit", "CR"],
        "balance_cols": ["Balance", "Closing Balance"],
        "date_formats": ["%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"],
    },
    "GENERIC": {
        "date_cols": ["Date", "Txn Date", "Transaction Date", "Value Date", "Posting Date"],
        "desc_cols": ["Description", "Narration", "Remarks", "Details", "Particulars", "Memo"],
        "ref_cols": ["Reference", "Ref No", "Cheque No", "Chq/Ref No", "UTR", "Transaction ID"],
        "debit_cols": ["Debit", "Debit Amount", "Withdrawal", "DR", "Debit(INR)"],
        "credit_cols": ["Credit", "Credit Amount", "Deposit", "CR", "Credit(INR)"],
        "balance_cols": ["Balance", "Closing Balance", "Running Balance", "Balance(INR)"],
        "date_formats": ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%m/%d/%Y"],
    },
}


def _find_column(headers: List[str], candidates: List[str]) -> Optional[str]:
    """Find the first matching column name (case-insensitive)."""
    header_lower = {h.strip().lower(): h for h in headers}
    for candidate in candidates:
        if candidate.strip().lower() in header_lower:
            return header_lower[candidate.strip().lower()]
    return None


def _parse_date_value(val: str, formats: List[str]) -> Optional[date]:
    """Try multiple date formats."""
    if not val or not val.strip():
        return None
    val = val.strip()
    for fmt in formats:
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            continue
    return None


def _parse_decimal(val: str) -> Decimal:
    """Parse a decimal value from string, handling commas and negatives."""
    if not val or not val.strip():
        return Decimal("0")
    val = val.strip().replace(",", "").replace(" ", "")
    # Handle negative in parentheses: (1,234.56) -> -1234.56
    if val.startswith("(") and val.endswith(")"):
        val = "-" + val[1:-1]
    try:
        return Decimal(val)
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _detect_bank_format(headers: List[str]) -> str:
    """Auto-detect bank format from column headers."""
    for bank_name, fmt in BANK_FORMATS.items():
        if bank_name == "GENERIC":
            continue
        matches = 0
        for candidate_list in fmt.values():
            if isinstance(candidate_list, list):
                for candidate in candidate_list:
                    if any(candidate.strip().lower() == h.strip().lower() for h in headers):
                        matches += 1
                        break
        if matches >= 3:
            return bank_name
    return "GENERIC"


def parse_csv_statement(file_content: bytes, bank_format: str = "AUTO") -> List[dict]:
    """Parse a CSV bank statement and return normalized transactions."""
    # Try different encodings
    for encoding in ["utf-8", "latin-1", "cp1252", "iso-8859-1"]:
        try:
            text = file_content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Could not decode file. Please use UTF-8 or Latin-1 encoding.")

    # Detect delimiter
    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(text[:4096])
    except csv.Error:
        dialect = csv.excel

    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    headers = reader.fieldnames
    if not headers:
        raise ValueError("Could not detect column headers in CSV file.")

    # Auto-detect format
    if bank_format == "AUTO":
        bank_format = _detect_bank_format(headers)

    fmt = BANK_FORMATS.get(bank_format, BANK_FORMATS["GENERIC"])

    # Find column mappings
    date_col = _find_column(headers, fmt["date_cols"])
    desc_col = _find_column(headers, fmt["desc_cols"])
    ref_col = _find_column(headers, fmt["ref_cols"])
    debit_col = _find_column(headers, fmt["debit_cols"])
    credit_col = _find_column(headers, fmt["credit_cols"])
    balance_col = _find_column(headers, fmt["balance_cols"])

    if not date_col:
        raise ValueError(f"Could not find a date column. Headers found: {headers}")
    if not debit_col and not credit_col:
        # Try single amount column
        amount_col = _find_column(headers, ["Amount", "Transaction Amount", "Value"])
        if not amount_col:
            raise ValueError(f"Could not find debit/credit columns. Headers found: {headers}")

    transactions = []
    for row_num, row in enumerate(reader, start=2):
        try:
            txn_date = _parse_date_value(row.get(date_col, ""), fmt["date_formats"])
            if not txn_date:
                continue

            description = (row.get(desc_col, "") or "").strip()
            reference = (row.get(ref_col, "") or "").strip()

            if debit_col and credit_col:
                debit = _parse_decimal(row.get(debit_col, "0"))
                credit = _parse_decimal(row.get(credit_col, "0"))
            else:
                # Single amount column - positive = credit, negative = debit
                amount = _parse_decimal(row.get(amount_col, "0"))
                if amount >= 0:
                    debit = Decimal("0")
                    credit = amount
                else:
                    debit = abs(amount)
                    credit = Decimal("0")

            balance = _parse_decimal(row.get(balance_col, "0")) if balance_col else None

            # Skip zero-amount rows
            if debit == 0 and credit == 0:
                continue

            # Amount: positive = credit (deposit), negative = debit (withdrawal)
            amount = credit - debit

            transactions.append({
                "transaction_date": txn_date,
                "amount": amount,
                "description": description[:500] if description else None,
                "reference_number": reference[:50] if reference else None,
                "balance": balance,
            })
        except Exception:
            continue  # Skip malformed rows

    if not transactions:
        raise ValueError("No valid transactions found in the file. Check the format.")

    return transactions


def parse_excel_statement(file_content: bytes, bank_format: str = "AUTO") -> List[dict]:
    """Parse an Excel bank statement."""
    try:
        import openpyxl
    except ImportError:
        raise ValueError("Excel parsing requires openpyxl. Install with: pip install openpyxl")

    wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    ws = wb.active

    # Read headers from first row
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(str(cell.value).strip() if cell.value else "")

    if bank_format == "AUTO":
        bank_format = _detect_bank_format(headers)

    fmt = BANK_FORMATS.get(bank_format, BANK_FORMATS["GENERIC"])

    date_col_idx = None
    desc_col_idx = None
    ref_col_idx = None
    debit_col_idx = None
    credit_col_idx = None
    balance_col_idx = None
    amount_col_idx = None

    for i, h in enumerate(headers):
        h_lower = h.strip().lower()
        if date_col_idx is None and any(c.lower() == h_lower for c in fmt["date_cols"]):
            date_col_idx = i
        if desc_col_idx is None and any(c.lower() == h_lower for c in fmt["desc_cols"]):
            desc_col_idx = i
        if ref_col_idx is None and any(c.lower() == h_lower for c in fmt["ref_cols"]):
            ref_col_idx = i
        if debit_col_idx is None and any(c.lower() == h_lower for c in fmt["debit_cols"]):
            debit_col_idx = i
        if credit_col_idx is None and any(c.lower() == h_lower for c in fmt["credit_cols"]):
            credit_col_idx = i
        if balance_col_idx is None and any(c.lower() == h_lower for c in fmt["balance_cols"]):
            balance_col_idx = i

    if date_col_idx is None:
        raise ValueError(f"Could not find a date column. Headers: {headers}")
    if debit_col_idx is None and credit_col_idx is None:
        for i, h in enumerate(headers):
            if h.lower() in ("amount", "transaction amount", "value"):
                amount_col_idx = i
                break
        if amount_col_idx is None:
            raise ValueError(f"Could not find debit/credit columns. Headers: {headers}")

    transactions = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            raw_date = row[date_col_idx]
            if raw_date is None:
                continue
            if isinstance(raw_date, datetime):
                txn_date = raw_date.date()
            elif isinstance(raw_date, date):
                txn_date = raw_date
            else:
                txn_date = _parse_date_value(str(raw_date), fmt["date_formats"])
            if not txn_date:
                continue

            description = str(row[desc_col_idx] or "").strip() if desc_col_idx is not None else ""
            reference = str(row[ref_col_idx] or "").strip() if ref_col_idx is not None else ""

            if debit_col_idx is not None and credit_col_idx is not None:
                debit = _parse_decimal(str(row[debit_col_idx] or "0"))
                credit = _parse_decimal(str(row[credit_col_idx] or "0"))
            else:
                amount = _parse_decimal(str(row[amount_col_idx] or "0"))
                if amount >= 0:
                    debit = Decimal("0")
                    credit = amount
                else:
                    debit = abs(amount)
                    credit = Decimal("0")

            balance = _parse_decimal(str(row[balance_col_idx] or "0")) if balance_col_idx is not None else None

            if debit == 0 and credit == 0:
                continue

            amount = credit - debit

            transactions.append({
                "transaction_date": txn_date,
                "amount": amount,
                "description": description[:500] if description else None,
                "reference_number": reference[:50] if reference else None,
                "balance": balance,
            })
        except Exception:
            continue

    wb.close()

    if not transactions:
        raise ValueError("No valid transactions found in the Excel file.")

    return transactions


# ---------------------------------------------------------------------------
# FILE UPLOAD ENDPOINT
# ---------------------------------------------------------------------------

@router.post("/upload", status_code=status.HTTP_201_CREATED)
async def upload_bank_statement(
    banking_profile_id: uuid.UUID = Query(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("invoice:create")),
):
    """Upload a bank statement CSV/Excel file. Parses transactions and creates statement."""
    # Verify banking profile
    profile = db.query(BankingProfile).filter(
        BankingProfile.id == banking_profile_id,
        BankingProfile.tenant_id == tenant_id,
    ).first()
    if not profile:
        raise HTTPException(status_code=404, detail="Banking profile not found.")

    # Read file
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file uploaded.")

    filename = (file.filename or "").lower()

    # Parse based on file type
    try:
        if filename.endswith(".csv") or filename.endswith(".txt"):
            transactions = parse_csv_statement(content)
        elif filename.endswith(".xlsx") or filename.endswith(".xls"):
            transactions = parse_excel_statement(content)
        else:
            # Try CSV first, then Excel
            try:
                transactions = parse_csv_statement(content)
            except Exception:
                transactions = parse_excel_statement(content)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")

    if not transactions:
        raise HTTPException(status_code=400, detail="No transactions found in file.")

    # Compute balances
    dates = [t["transaction_date"] for t in transactions]
    amounts = [t["amount"] for t in transactions]
    starting_balance = Decimal("0")
    ending_balance = sum(amounts)

    # Check for duplicate import (same file, same date range)
    existing = db.query(BankStatement).filter(
        BankStatement.tenant_id == tenant_id,
        BankStatement.banking_profile_id == banking_profile_id,
        BankStatement.statement_date == min(dates),
    ).first()
    if existing:
        raise HTTPException(
            status_code=400,
            detail=f"A statement for this date ({min(dates)}) already exists. Delete it first or use a different date range."
        )

    # Create statement
    statement = BankStatement(
        tenant_id=tenant_id,
        banking_profile_id=banking_profile_id,
        statement_date=min(dates),
        starting_balance=starting_balance,
        ending_balance=ending_balance,
        currency="INR",
        status="IMPORTED",
    )
    db.add(statement)
    db.flush()

    # Create transactions
    for txn_data in transactions:
        txn = BankTransaction(
            bank_statement_id=statement.id,
            transaction_date=txn_data["transaction_date"],
            amount=txn_data["amount"],
            description=txn_data["description"],
            reference_number=txn_data["reference_number"],
            status="PENDING",
        )
        db.add(txn)

    db.commit()
    db.refresh(statement)

    return {
        "statement_id": str(statement.id),
        "transactions_imported": len(transactions),
        "date_range": f"{min(dates)} to {max(dates)}",
        "total_credits": str(sum(t["amount"] for t in transactions if t["amount"] > 0)),
        "total_debits": str(abs(sum(t["amount"] for t in transactions if t["amount"] < 0))),
    }


# ---------------------------------------------------------------------------
# STATEMENT ENDPOINTS
# ---------------------------------------------------------------------------

@router.get("/statements", response_model=List[BankStatementListResponse])
def list_bank_statements(
    page: int = 1,
    limit: int = 50,
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("invoice:view")),
):
    offset = (page - 1) * limit
    results = db.query(BankStatement).filter(
        BankStatement.tenant_id == tenant_id
    ).order_by(BankStatement.created_at.desc()).offset(offset).limit(limit).all()

    response = []
    for stmt in results:
        response.append(BankStatementListResponse(
            id=stmt.id,
            banking_profile_id=stmt.banking_profile_id,
            bank_name=stmt.banking_profile.bank_name if stmt.banking_profile else None,
            account_number=stmt.banking_profile.account_number if stmt.banking_profile else None,
            statement_date=stmt.statement_date,
            starting_balance=stmt.starting_balance,
            ending_balance=stmt.ending_balance,
            closing_balance=stmt.ending_balance,
            currency=stmt.currency,
            status=stmt.status,
            created_at=stmt.created_at
        ))
    return response


@router.get("/statements/{id}", response_model=BankStatementResponse)
def get_bank_statement(
    id: uuid.UUID,
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("invoice:view")),
):
    statement = db.query(BankStatement).filter(
        BankStatement.id == id,
        BankStatement.tenant_id == tenant_id
    ).first()
    if not statement:
        raise HTTPException(status_code=404, detail="Bank statement not found.")
    return statement


@router.delete("/statements/{id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_bank_statement(
    id: uuid.UUID,
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("invoice:delete")),
):
    """Delete a bank statement and all its transactions."""
    statement = db.query(BankStatement).filter(
        BankStatement.id == id,
        BankStatement.tenant_id == tenant_id,
    ).first()
    if not statement:
        raise HTTPException(status_code=404, detail="Bank statement not found.")

    # Check for reconciled transactions
    reconciled = db.query(BankTransaction).filter(
        BankTransaction.bank_statement_id == id,
        BankTransaction.status == "RECONCILED",
    ).count()
    if reconciled > 0:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete statement with {reconciled} reconciled transactions. Undo reconciliations first."
        )

    db.delete(statement)
    db.commit()


# ---------------------------------------------------------------------------
# STATEMENT STATS
# ---------------------------------------------------------------------------

@router.get("/statements/{id}/stats")
def get_statement_stats(
    id: uuid.UUID,
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("invoice:view")),
):
    """Get reconciliation statistics for a statement."""
    statement = db.query(BankStatement).filter(
        BankStatement.id == id,
        BankStatement.tenant_id == tenant_id,
    ).first()
    if not statement:
        raise HTTPException(status_code=404, detail="Bank statement not found.")

    total = db.query(func.count(BankTransaction.id)).filter(
        BankTransaction.bank_statement_id == id,
    ).scalar() or 0

    reconciled = db.query(func.count(BankTransaction.id)).filter(
        BankTransaction.bank_statement_id == id,
        BankTransaction.status == "RECONCILED",
    ).scalar() or 0

    credits = db.query(func.coalesce(func.sum(BankTransaction.amount), 0)).filter(
        BankTransaction.bank_statement_id == id,
        BankTransaction.amount > 0,
    ).scalar() or 0

    debits = db.query(func.coalesce(func.sum(func.abs(BankTransaction.amount)), 0)).filter(
        BankTransaction.bank_statement_id == id,
        BankTransaction.amount < 0,
    ).scalar() or 0

    return StatementStats(
        total_transactions=total,
        reconciled=reconciled,
        pending=total - reconciled,
        total_credits=Decimal(str(credits)),
        total_debits=Decimal(str(debits)),
        reconciliation_pct=round(reconciled / total * 100, 1) if total > 0 else 0.0,
    )


# ---------------------------------------------------------------------------
# TRANSACTION ENDPOINTS
# ---------------------------------------------------------------------------

@router.get("/statements/{statement_id}/transactions", response_model=List[BankTransactionResponse])
def list_bank_transactions(
    statement_id: uuid.UUID,
    status_filter: Optional[str] = None,
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("invoice:view")),
):
    statement = db.query(BankStatement).filter(
        BankStatement.id == statement_id,
        BankStatement.tenant_id == tenant_id
    ).first()
    if not statement:
        raise HTTPException(status_code=404, detail="Bank statement not found.")

    q = db.query(BankTransaction).filter(BankTransaction.bank_statement_id == statement_id)
    if status_filter:
        q = q.filter(BankTransaction.status == status_filter.upper())

    transactions = q.order_by(BankTransaction.transaction_date.asc()).all()

    response = []
    for txn in transactions:
        response.append(BankTransactionResponse(
            id=txn.id,
            transaction_date=txn.transaction_date,
            amount=txn.amount,
            description=txn.description,
            reference_number=txn.reference_number,
            status=txn.status,
            created_at=txn.created_at,
            updated_at=txn.updated_at
        ))
    return response


# ---------------------------------------------------------------------------
# AUTO-MATCH WITH DEDUPLICATION
# ---------------------------------------------------------------------------

@router.post("/statements/{statement_id}/auto-match")
def auto_match_transactions(
    statement_id: uuid.UUID,
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("invoice:finalize")),
):
    """Auto-match unreconciled bank transactions to payments and bills.
    Uses scoring algorithm with deduplication (each payment matched at most once)."""
    from difflib import SequenceMatcher

    statement = db.query(BankStatement).filter(
        BankStatement.id == statement_id,
        BankStatement.tenant_id == tenant_id
    ).first()
    if not statement:
        raise HTTPException(status_code=404, detail="Bank statement not found.")

    unmatched_txns = db.query(BankTransaction).filter(
        BankTransaction.bank_statement_id == statement_id,
        BankTransaction.status == "PENDING"
    ).order_by(BankTransaction.transaction_date.asc()).all()

    if not unmatched_txns:
        return {"matched": 0, "matches": [], "message": "No pending transactions to match."}

    # Load candidates - only those not already reconciled
    already_reconciled_payment_ids = set(
        r[0] for r in db.query(BankReconciliation.payment_id).filter(
            BankReconciliation.tenant_id == tenant_id,
            BankReconciliation.payment_id != None,
        ).all()
    )
    already_reconciled_bp_ids = set(
        r[0] for r in db.query(BankReconciliation.bill_payment_id).filter(
            BankReconciliation.tenant_id == tenant_id,
            BankReconciliation.bill_payment_id != None,
        ).all()
    )

    payments = db.query(Payment).filter(
        Payment.tenant_id == tenant_id,
        Payment.status == "ACTIVE",
        Payment.deleted_at == None,
        ~Payment.id.in_(already_reconciled_payment_ids) if already_reconciled_payment_ids else True,
    ).all()

    bill_payments = db.query(BillPayment).filter(
        BillPayment.tenant_id == tenant_id,
        BillPayment.status == "ACTIVE",
        BillPayment.deleted_at == None,
        ~BillPayment.id.in_(already_reconciled_bp_ids) if already_reconciled_bp_ids else True,
    ).all()

    # Build match candidates with pre-computed keys
    payment_candidates = []
    for pmt in payments:
        payment_candidates.append({
            "type": "payment",
            "id": pmt.id,
            "amount": pmt.amount,
            "date": pmt.payment_date,
            "ref": (pmt.reference_number or "").lower(),
        })

    bp_candidates = []
    for bp in bill_payments:
        bp_candidates.append({
            "type": "bill_payment",
            "id": bp.id,
            "amount": bp.amount,
            "date": bp.payment_date,
            "ref": (bp.reference_number or "").lower(),
        })

    all_candidates = payment_candidates + bp_candidates
    matched_candidate_ids = set()
    matches = []

    for txn in unmatched_txns:
        best_match = None
        best_score = 0

        txn_ref = (txn.reference_number or "").lower()
        txn_date = txn.transaction_date
        txn_amount = txn.amount

        for cand in all_candidates:
            # Skip already matched candidates
            if cand["id"] in matched_candidate_ids:
                continue

            score = 0

            # Amount match (40 pts)
            cand_amount = cand["amount"]
            amount_diff = abs(abs(txn_amount) - cand_amount)
            if amount_diff < Decimal("0.01"):
                score += 40
            elif amount_diff < Decimal("1.00"):
                score += 30
            elif amount_diff / max(cand_amount, Decimal("1")) < Decimal("0.05"):
                score += 20

            # Date proximity (30 pts)
            if txn_date and cand["date"]:
                days_diff = abs((txn_date - cand["date"]).days)
                if days_diff <= 1:
                    score += 30
                elif days_diff <= 3:
                    score += 25
                elif days_diff <= 7:
                    score += 15
                elif days_diff <= 14:
                    score += 5

            # Reference similarity (30 pts)
            if txn_ref and cand["ref"]:
                ratio = SequenceMatcher(None, txn_ref, cand["ref"]).ratio()
                score += int(ratio * 30)

            if score > best_score and score >= 60:
                best_score = score
                best_match = cand

        if best_match:
            matched_candidate_ids.add(best_match["id"])
            recon = BankReconciliation(
                tenant_id=tenant_id,
                bank_transaction_id=txn.id,
                payment_id=best_match["id"] if best_match["type"] == "payment" else None,
                bill_payment_id=best_match["id"] if best_match["type"] == "bill_payment" else None,
                amount=best_match["amount"],
                notes=f"Auto-matched (score: {best_score})",
            )
            txn.status = "RECONCILED"
            db.add(recon)
            matches.append({
                "transaction_id": str(txn.id),
                "matched_type": best_match["type"],
                "matched_id": str(best_match["id"]),
                "amount": str(best_match["amount"]),
                "score": best_score,
            })

    db.commit()
    return {"matched": len(matches), "matches": matches}


# ---------------------------------------------------------------------------
# MATCH SUGGESTIONS (without auto-committing)
# ---------------------------------------------------------------------------

@router.get("/statements/{statement_id}/suggestions")
def get_match_suggestions(
    statement_id: uuid.UUID,
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("invoice:view")),
):
    """Get match suggestions for pending transactions without committing."""
    from difflib import SequenceMatcher

    statement = db.query(BankStatement).filter(
        BankStatement.id == statement_id,
        BankStatement.tenant_id == tenant_id
    ).first()
    if not statement:
        raise HTTPException(status_code=404, detail="Bank statement not found.")

    pending_txns = db.query(BankTransaction).filter(
        BankTransaction.bank_statement_id == statement_id,
        BankTransaction.status == "PENDING"
    ).order_by(BankTransaction.transaction_date.asc()).all()

    # Load all candidates
    payments = db.query(Payment).filter(
        Payment.tenant_id == tenant_id,
        Payment.status == "ACTIVE",
        Payment.deleted_at == None,
    ).all()

    bill_payments = db.query(BillPayment).filter(
        BillPayment.tenant_id == tenant_id,
        BillPayment.status == "ACTIVE",
        BillPayment.deleted_at == None,
    ).all()

    suggestions = []
    for txn in pending_txns:
        txn_ref = (txn.reference_number or "").lower()
        txn_date = txn.transaction_date
        txn_amount = txn.amount

        candidates = []

        for pmt in payments:
            score = _calculate_match_score(txn_amount, txn_date, txn_ref, pmt.amount, pmt.payment_date, (pmt.reference_number or "").lower())
            if score >= 40:
                candidates.append({
                    "type": "payment",
                    "id": str(pmt.id),
                    "amount": str(pmt.amount),
                    "date": str(pmt.payment_date),
                    "reference": pmt.reference_number,
                    "score": score,
                })

        for bp in bill_payments:
            score = _calculate_match_score(abs(txn_amount), txn_date, txn_ref, bp.amount, bp.payment_date, (bp.reference_number or "").lower())
            if score >= 40:
                candidates.append({
                    "type": "bill_payment",
                    "id": str(bp.id),
                    "amount": str(bp.amount),
                    "date": str(bp.payment_date),
                    "reference": bp.reference_number,
                    "score": score,
                })

        # Sort by score descending, take top 3
        candidates.sort(key=lambda x: x["score"], reverse=True)
        candidates = candidates[:3]

        suggestions.append(MatchSuggestion(
            transaction_id=str(txn.id),
            transaction_date=str(txn.transaction_date),
            transaction_amount=str(txn.amount),
            transaction_description=txn.description,
            suggested_matches=candidates,
        ))

    return suggestions


def _calculate_match_score(
    txn_amount: Decimal, txn_date: date, txn_ref: str,
    cand_amount: Decimal, cand_date: date, cand_ref: str,
) -> int:
    """Calculate match score between transaction and candidate."""
    from difflib import SequenceMatcher

    score = 0

    # Amount (40 pts)
    amount_diff = abs(abs(txn_amount) - cand_amount)
    if amount_diff < Decimal("0.01"):
        score += 40
    elif amount_diff < Decimal("1.00"):
        score += 30
    elif amount_diff / max(cand_amount, Decimal("1")) < Decimal("0.05"):
        score += 20

    # Date (30 pts)
    if txn_date and cand_date:
        days_diff = abs((txn_date - cand_date).days)
        if days_diff <= 1:
            score += 30
        elif days_diff <= 3:
            score += 25
        elif days_diff <= 7:
            score += 15
        elif days_diff <= 14:
            score += 5

    # Reference (30 pts)
    if txn_ref and cand_ref:
        ratio = SequenceMatcher(None, txn_ref, cand_ref).ratio()
        score += int(ratio * 30)

    return score


# ---------------------------------------------------------------------------
# BULK RECONCILE
# ---------------------------------------------------------------------------

@router.post("/bulk-reconcile")
def bulk_reconcile(
    payload: BulkReconcileRequest,
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("invoice:finalize")),
):
    """Bulk reconcile multiple transactions at once."""
    if not payload.items:
        raise HTTPException(status_code=400, detail="No items to reconcile.")

    reconciled = 0
    errors = []

    for item in payload.items:
        try:
            # Verify transaction
            txn = db.query(BankTransaction).join(BankStatement).filter(
                BankTransaction.id == item.transaction_id,
                BankStatement.tenant_id == tenant_id,
            ).first()
            if not txn:
                errors.append(f"Transaction {item.transaction_id} not found.")
                continue
            if txn.status != "PENDING":
                errors.append(f"Transaction {item.transaction_id} is not pending.")
                continue

            # Verify payment
            if item.payment_id:
                payment = db.query(Payment).filter(
                    Payment.id == item.payment_id,
                    Payment.tenant_id == tenant_id,
                ).first()
                if not payment:
                    errors.append(f"Payment {item.payment_id} not found.")
                    continue

            if item.bill_payment_id:
                bp = db.query(BillPayment).filter(
                    BillPayment.id == item.bill_payment_id,
                    BillPayment.tenant_id == tenant_id,
                ).first()
                if not bp:
                    errors.append(f"Bill payment {item.bill_payment_id} not found.")
                    continue

            # Check not already reconciled
            existing = db.query(BankReconciliation).filter(
                BankReconciliation.bank_transaction_id == item.transaction_id,
            ).first()
            if existing:
                errors.append(f"Transaction {item.transaction_id} already reconciled.")
                continue

            recon = BankReconciliation(
                tenant_id=tenant_id,
                bank_transaction_id=item.transaction_id,
                payment_id=item.payment_id,
                bill_payment_id=item.bill_payment_id,
                amount=item.amount,
                notes="Bulk reconciled",
            )
            txn.status = "RECONCILED"
            db.add(recon)
            reconciled += 1

        except Exception as e:
            errors.append(f"Error processing {item.transaction_id}: {str(e)}")

    db.commit()
    return {"reconciled": reconciled, "errors": errors}


# ---------------------------------------------------------------------------
# PENDING INVOICES & BILLS (for manual matching UI)
# ---------------------------------------------------------------------------

@router.get("/pending-invoices")
def get_pending_invoices(
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("invoice:view")),
):
    """Get outstanding invoices that could be matched to bank transactions."""
    invoices = db.query(Invoice).join(Contact, Invoice.contact_id == Contact.id).filter(
        Invoice.tenant_id == tenant_id,
        Invoice.status.in_(["POSTED", "PARTIALLY_PAID"]),
        Invoice.deleted_at == None,
    ).order_by(Invoice.due_date.asc()).all()

    today = date.today()
    result = []
    for inv in invoices:
        outstanding = inv.total - (inv.amount_paid or Decimal("0"))
        if outstanding <= 0:
            continue
        days_overdue = max(0, (today - inv.due_date).days) if inv.due_date else 0
        result.append(PendingInvoice(
            id=str(inv.id),
            invoice_number=inv.invoice_number,
            contact_name=inv.contact.name if inv.contact else None,
            issue_date=str(inv.issue_date),
            due_date=str(inv.due_date),
            total=inv.total,
            amount_paid=inv.amount_paid or Decimal("0"),
            outstanding=outstanding,
            days_overdue=days_overdue,
        ))

    return result


@router.get("/pending-bills")
def get_pending_bills(
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("invoice:view")),
):
    """Get outstanding bills that could be matched to bank transactions."""
    bills = db.query(Bill).join(Contact, Bill.contact_id == Contact.id).filter(
        Bill.tenant_id == tenant_id,
        Bill.status.in_(["POSTED", "PARTIALLY_PAID"]),
        Bill.deleted_at == None,
    ).order_by(Bill.due_date.asc()).all()

    today = date.today()
    result = []
    for bill in bills:
        outstanding = bill.total - (bill.amount_paid or Decimal("0"))
        if outstanding <= 0:
            continue
        days_overdue = max(0, (today - bill.due_date).days) if bill.due_date else 0
        result.append(PendingBill(
            id=str(bill.id),
            bill_number=bill.bill_number,
            contact_name=bill.contact.name if bill.contact else None,
            issue_date=str(bill.issue_date),
            due_date=str(bill.due_date),
            total=bill.total,
            amount_paid=bill.amount_paid or Decimal("0"),
            outstanding=outstanding,
            days_overdue=days_overdue,
        ))

    return result


# ---------------------------------------------------------------------------
# RECONCILIATION ENDPOINTS
# ---------------------------------------------------------------------------

@router.post("/transactions/{transaction_id}/reconcile", response_model=BankReconciliationResponse, status_code=status.HTTP_201_CREATED)
def reconcile_bank_transaction(
    transaction_id: uuid.UUID,
    payload: BankReconciliationCreate,
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("invoice:finalize")),
):
    transaction = db.query(BankTransaction).join(BankStatement).filter(
        BankTransaction.id == transaction_id,
        BankStatement.tenant_id == tenant_id
    ).first()
    if not transaction:
        raise HTTPException(status_code=404, detail="Bank transaction not found.")

    if transaction.status != "PENDING":
        raise HTTPException(status_code=400, detail="Only pending transactions can be reconciled.")

    if payload.payment_id is not None and payload.bill_payment_id is not None:
        raise HTTPException(status_code=400, detail="Cannot reconcile with both payment and bill payment.")

    if payload.payment_id is None and payload.bill_payment_id is None:
        raise HTTPException(status_code=400, detail="Must specify either payment_id or bill_payment_id.")

    if payload.payment_id:
        payment = db.query(Payment).filter(
            Payment.id == payload.payment_id,
            Payment.tenant_id == tenant_id
        ).first()
        if not payment:
            raise HTTPException(status_code=404, detail="Payment not found.")

    if payload.bill_payment_id:
        bill_payment = db.query(BillPayment).filter(
            BillPayment.id == payload.bill_payment_id,
            BillPayment.tenant_id == tenant_id
        ).first()
        if not bill_payment:
            raise HTTPException(status_code=404, detail="Bill payment not found.")

    existing = db.query(BankReconciliation).filter(
        BankReconciliation.bank_transaction_id == transaction_id
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Transaction already reconciled.")

    reconciliation = BankReconciliation(
        tenant_id=tenant_id,
        bank_transaction_id=transaction_id,
        payment_id=payload.payment_id,
        bill_payment_id=payload.bill_payment_id,
        amount=payload.amount,
        notes=payload.notes
    )

    transaction.status = "RECONCILED"

    db.add(reconciliation)
    db.commit()
    db.refresh(reconciliation)
    return reconciliation


@router.get("/reconciliations", response_model=List[BankReconciliationListResponse])
def list_bank_reconciliations(
    page: int = 1,
    limit: int = 50,
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("invoice:view")),
):
    offset = (page - 1) * limit
    results = db.query(BankReconciliation).join(BankTransaction).join(BankStatement).filter(
        BankStatement.tenant_id == tenant_id
    ).offset(offset).limit(limit).all()

    response = []
    for recon in results:
        response.append(BankReconciliationListResponse(
            id=recon.id,
            bank_transaction_id=recon.bank_transaction_id,
            amount=recon.amount,
            notes=recon.notes,
            created_at=recon.created_at
        ))
    return response


@router.get("/reconciliations/{id}", response_model=BankReconciliationResponse)
def get_bank_reconciliation(
    id: uuid.UUID,
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("invoice:view")),
):
    reconciliation = db.query(BankReconciliation).join(BankTransaction).join(BankStatement).filter(
        BankReconciliation.id == id,
        BankStatement.tenant_id == tenant_id
    ).first()
    if not reconciliation:
        raise HTTPException(status_code=404, detail="Bank reconciliation not found.")
    return reconciliation


@router.post("/reconciliations/{id}/undo", response_model=BankReconciliationResponse)
def undo_bank_reconciliation(
    id: uuid.UUID,
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("invoice:finalize")),
):
    reconciliation = db.query(BankReconciliation).join(BankTransaction).join(BankStatement).filter(
        BankReconciliation.id == id,
        BankStatement.tenant_id == tenant_id
    ).first()
    if not reconciliation:
        raise HTTPException(status_code=404, detail="Bank reconciliation not found.")

    transaction = db.query(BankTransaction).filter(
        BankTransaction.id == reconciliation.bank_transaction_id
    ).first()
    if transaction:
        transaction.status = "PENDING"

    db.delete(reconciliation)
    db.commit()

    return reconciliation
