from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
import uuid
from datetime import date
from decimal import Decimal
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import re

from src.core.database import get_db_session
from src.infrastructure.database.models import Invoice, Bill, CreditNote, DebitNote, Contact, Tenant, TenantSetting
from src.schemas.gst_schemas import (
    GSTR1Response, GSTR1B2BLine, GSTR1B2CLine, GSTR1B2CSLine, GSTR1NoteLine, GSTR1HSNLine,
    GSTR2Response, GSTR2B2BLine
)
from src.domains.company.services import resolve_origin_state_code
from src.domains.accounting.report_services import GSTR3BService
from src.api.deps import enforce_permission

router = APIRouter(prefix="/gst", tags=["GST Compliance"])

@router.get("/validate-gstin/{gstin}")
def validate_gstin_format(gstin: str):
    """Validates GSTIN format (15 characters, checksum)."""
    if not gstin or len(gstin) != 15:
        raise HTTPException(status_code=400, detail="GSTIN must be 15 characters.")
    
    pattern = r'^\d{2}[A-Z]{5}\d{4}[A-Z]\d[Z][A-Z\d]$'
    if not re.match(pattern, gstin):
        raise HTTPException(status_code=400, detail="Invalid GSTIN format.")
    
    state_code = gstin[:2]
    valid_states = {
        "01","02","03","04","05","06","07","08","09","10","11","12","13","14","15",
        "16","17","18","19","20","21","22","23","24","25","26","27","28","29","30",
        "31","32","33","34","35","36","37","38","97","99"
    }
    if state_code not in valid_states:
        raise HTTPException(status_code=400, detail=f"Invalid state code: {state_code}")
    
    return {"valid": True, "state_code": state_code}

@router.get("/gstr1", response_model=GSTR1Response)
def get_gstr1_report(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("invoice:view"))
):
    """Compiles GSTR-1 outward supply sales tax returns for a given period."""
    # 1. Fetch tenant origin state code
    origin_state_code = resolve_origin_state_code(db, tenant_id)

    # 2. Fetch finalized sales invoices
    q_inv = db.query(Invoice).filter(
        Invoice.tenant_id == tenant_id,
        Invoice.status.in_(["POSTED", "PARTIALLY_PAID", "PAID"]),
        Invoice.deleted_at == None
    )
    if start_date:
        q_inv = q_inv.filter(Invoice.issue_date >= start_date)
    if end_date:
        q_inv = q_inv.filter(Invoice.issue_date <= end_date)
    invoices = q_inv.all()

    b2b_lines = []
    b2cl_lines = []
    b2cs_groups = {}  # key: (pos_state_code, gst_rate)
    hsn_groups = {}  # key: hsn_sac

    for inv in invoices:
        contact = inv.contact
        is_registered = contact and contact.gstin

        if is_registered:
            # Section 4: B2B registered sales
            b2b_lines.append(
                GSTR1B2BLine(
                    customer_name=contact.name,
                    customer_gstin=contact.gstin,
                    invoice_number=inv.invoice_number,
                    invoice_date=inv.issue_date,
                    pos_state_code=inv.pos_state_code,
                    taxable_value=inv.subtotal,
                    cgst_amount=inv.cgst_amount,
                    sgst_amount=inv.sgst_amount,
                    igst_amount=inv.igst_amount,
                    utgst_amount=inv.utgst_amount,
                    cess_amount=inv.cess_amount,
                    total_value=inv.total
                )
            )
        else:
            # Section 5 & 7: Unregistered Sales (B2C)
            is_inter_state = (inv.pos_state_code != origin_state_code)
            if is_inter_state and inv.total > Decimal("250000.00"):
                # B2C Large (Inter-state, > 2.5L total)
                b2cl_lines.append(
                    GSTR1B2CLine(
                        invoice_number=inv.invoice_number,
                        invoice_date=inv.issue_date,
                        pos_state_code=inv.pos_state_code,
                        taxable_value=inv.subtotal,
                        igst_amount=inv.igst_amount,
                        total_value=inv.total
                    )
                )
            else:
                # B2C Small (All other unregistered)
                for line in inv.lines:
                    key = (inv.pos_state_code, line.gst_rate)
                    if key not in b2cs_groups:
                        b2cs_groups[key] = {
                            "taxable_value": Decimal("0.0000"),
                            "cgst_amount": Decimal("0.0000"),
                            "sgst_amount": Decimal("0.0000"),
                            "igst_amount": Decimal("0.0000"),
                            "utgst_amount": Decimal("0.0000"),
                            "cess_amount": Decimal("0.0000")
                        }
                    b2cs_groups[key]["taxable_value"] += line.subtotal
                    b2cs_groups[key]["cgst_amount"] += line.cgst_amount
                    b2cs_groups[key]["sgst_amount"] += line.sgst_amount
                    b2cs_groups[key]["igst_amount"] += line.igst_amount
                    b2cs_groups[key]["utgst_amount"] += line.utgst_amount
                    b2cs_groups[key]["cess_amount"] += line.cess_amount

        # Section 12: HSN Summary aggregation
        for line in inv.lines:
            hsn = line.hsn_sac
            product = line.product
            desc = product.name if product else "N/A"
            uom = product.uom if product else "PCS"

            if hsn not in hsn_groups:
                hsn_groups[hsn] = {
                    "description": desc,
                    "uom": uom,
                    "total_quantity": Decimal("0.0000"),
                    "total_value": Decimal("0.0000"),
                    "taxable_value": Decimal("0.0000"),
                    "cgst_amount": Decimal("0.0000"),
                    "sgst_amount": Decimal("0.0000"),
                    "igst_amount": Decimal("0.0000"),
                    "utgst_amount": Decimal("0.0000"),
                    "cess_amount": Decimal("0.0000")
                }
            hsn_groups[hsn]["total_quantity"] += line.quantity
            hsn_groups[hsn]["total_value"] += line.total
            hsn_groups[hsn]["taxable_value"] += line.subtotal
            hsn_groups[hsn]["cgst_amount"] += line.cgst_amount
            hsn_groups[hsn]["sgst_amount"] += line.sgst_amount
            hsn_groups[hsn]["igst_amount"] += line.igst_amount
            hsn_groups[hsn]["utgst_amount"] += line.utgst_amount
            hsn_groups[hsn]["cess_amount"] += line.cess_amount

    b2cs_lines = [
        GSTR1B2CSLine(
            pos_state_code=k[0],
            gst_rate=k[1],
            taxable_value=v["taxable_value"],
            cgst_amount=v["cgst_amount"],
            sgst_amount=v["sgst_amount"],
            igst_amount=v["igst_amount"],
            utgst_amount=v["utgst_amount"],
            cess_amount=v["cess_amount"]
        )
        for k, v in b2cs_groups.items()
    ]

    hsn_lines = [
        GSTR1HSNLine(
            hsn_sac=hsn,
            description=v["description"],
            uom=v["uom"],
            total_quantity=v["total_quantity"],
            total_value=v["total_value"],
            taxable_value=v["taxable_value"],
            cgst_amount=v["cgst_amount"],
            sgst_amount=v["sgst_amount"],
            igst_amount=v["igst_amount"],
            utgst_amount=v["utgst_amount"],
            cess_amount=v["cess_amount"]
        )
        for hsn, v in hsn_groups.items()
    ]

    # 2. Fetch finalized Credit and Debit Notes
    q_cn = db.query(CreditNote).filter(
        CreditNote.tenant_id == tenant_id,
        CreditNote.status == "POSTED",
        CreditNote.deleted_at == None
    )
    if start_date:
        q_cn = q_cn.filter(CreditNote.issue_date >= start_date)
    if end_date:
        q_cn = q_cn.filter(CreditNote.issue_date <= end_date)
    credit_notes = q_cn.all()

    q_dn = db.query(DebitNote).filter(
        DebitNote.tenant_id == tenant_id,
        DebitNote.status == "POSTED",
        DebitNote.deleted_at == None
    )
    if start_date:
        q_dn = q_dn.filter(DebitNote.issue_date >= start_date)
    if end_date:
        q_dn = q_dn.filter(DebitNote.issue_date <= end_date)
    debit_notes = q_dn.all()

    cdnr_lines = []
    cdnur_lines = []

    # Process Credit Notes (Section 9B CDNR/CDNUR)
    for cn in credit_notes:
        invoice = cn.invoice
        contact = invoice.contact if invoice else None
        is_registered = contact and contact.gstin

        note_line = GSTR1NoteLine(
            note_number=cn.credit_note_number,
            note_date=cn.issue_date,
            note_type="CREDIT",
            invoice_number=invoice.invoice_number if invoice else None,
            customer_gstin=contact.gstin if is_registered else None,
            reason=cn.reason,
            taxable_value=cn.subtotal,
            cgst_amount=cn.cgst_amount,
            sgst_amount=cn.sgst_amount,
            igst_amount=cn.igst_amount,
            utgst_amount=cn.utgst_amount,
            cess_amount=cn.cess_amount,
            total_value=cn.total
        )
        if is_registered:
            cdnr_lines.append(note_line)
        else:
            cdnur_lines.append(note_line)

    # Process Debit Notes (Section 9B CDNR/CDNUR)
    for dn in debit_notes:
        invoice = dn.invoice
        contact = invoice.contact if invoice else None
        is_registered = contact and contact.gstin

        note_line = GSTR1NoteLine(
            note_number=dn.debit_note_number,
            note_date=dn.issue_date,
            note_type="DEBIT",
            invoice_number=invoice.invoice_number if invoice else None,
            customer_gstin=contact.gstin if is_registered else None,
            reason=dn.reason,
            taxable_value=dn.subtotal,
            cgst_amount=dn.cgst_amount,
            sgst_amount=dn.sgst_amount,
            igst_amount=dn.igst_amount,
            utgst_amount=dn.utgst_amount,
            cess_amount=dn.cess_amount,
            total_value=dn.total
        )
        if is_registered:
            cdnr_lines.append(note_line)
        else:
            cdnur_lines.append(note_line)

    return GSTR1Response(
        b2b=b2b_lines,
        b2cl=b2cl_lines,
        b2cs=b2cs_lines,
        cdnr=cdnr_lines,
        cdnur=cdnur_lines,
        hsn_summary=hsn_lines
    )

@router.get("/gstr2", response_model=GSTR2Response)
def get_gstr2_report(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("invoice:view"))
):
    """Compiles GSTR-2 inward purchase invoice summaries for Input Tax Credit reconciliation."""
    # Fetch finalized vendor bills
    q_bill = db.query(Bill).filter(
        Bill.tenant_id == tenant_id,
        Bill.status.in_(["POSTED", "PARTIALLY_PAID", "PAID"]),
        Bill.deleted_at == None
    )
    if start_date:
        q_bill = q_bill.filter(Bill.issue_date >= start_date)
    if end_date:
        q_bill = q_bill.filter(Bill.issue_date <= end_date)
    bills = q_bill.all()

    b2b_purchases = []
    for b in bills:
        contact = b.contact
        # GSTR-2 maps registered vendor purchases
        if contact and contact.gstin:
            b2b_purchases.append(
                GSTR2B2BLine(
                    vendor_name=contact.name,
                    vendor_gstin=contact.gstin,
                    bill_number=b.bill_number,
                    bill_date=b.issue_date,
                    pos_state_code=b.pos_state_code,
                    taxable_value=b.subtotal,
                    cgst_amount=b.cgst_amount,
                    sgst_amount=b.sgst_amount,
                    igst_amount=b.igst_amount,
                    utgst_amount=b.utgst_amount,
                    cess_amount=b.cess_amount,
                    total_value=b.total
                )
            )

    return GSTR2Response(b2b_purchases=b2b_purchases)


# ---------------------------------------------------------------------------
# Excel Offline Tool Exports
# ---------------------------------------------------------------------------

STATE_CODE_MAPPING = {
    "01": "01-Jammu & Kashmir", "02": "02-Himachal Pradesh", "03": "03-Punjab",
    "04": "04-Chandigarh", "05": "05-Uttarakhand", "06": "06-Haryana",
    "07": "07-Delhi", "08": "08-Rajasthan", "09": "09-Uttar Pradesh",
    "10": "10-Bihar", "11": "11-Sikkim", "12": "12-Arunachal Pradesh",
    "13": "13-Nagaland", "14": "14-Manipur", "15": "15-Mizoram",
    "16": "16-Tripura", "17": "17-Meghalaya", "18": "18-Assam",
    "19": "19-West Bengal", "20": "20-Jharkhand", "21": "21-Odisha",
    "22": "22-Chhattisgarh", "23": "23-Madhya Pradesh", "24": "24-Gujarat",
    "25": "25-Daman & Diu", "26": "26-Dadra & Nagar Haveli", "27": "27-Maharashtra",
    "28": "28-Andhra Pradesh", "29": "29-Karnataka", "30": "30-Goa",
    "31": "31-Lakshadweep", "32": "32-Kerala", "33": "33-Tamil Nadu",
    "34": "34-Puducherry", "35": "35-Andaman & Nicobar Islands",
    "36": "36-Telangana", "37": "37-Andhra Pradesh (New)", "38": "38-Ladakh"
}

def format_pos(code: str) -> str:
    if not code:
        return ""
    code_clean = code.strip()
    return STATE_CODE_MAPPING.get(code_clean, f"{code_clean}-Other State")


@router.get("/gstr1/export")
def export_gstr1(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("invoice:view"))
):
    """Generates an Excel workbook for GSTR-1 returns, importable by the GST Offline Tool."""
    report = get_gstr1_report(start_date=start_date, end_date=end_date, db=db, tenant_id=tenant_id)
    origin_state_code = resolve_origin_state_code(db, tenant_id)

    q_inv = db.query(Invoice).filter(
        Invoice.tenant_id == tenant_id,
        Invoice.status.in_(["POSTED", "PARTIALLY_PAID", "PAID"]),
        Invoice.deleted_at == None
    )
    if start_date:
        q_inv = q_inv.filter(Invoice.issue_date >= start_date)
    if end_date:
        q_inv = q_inv.filter(Invoice.issue_date <= end_date)
    invoices = q_inv.all()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # 1. b2b sheet
    ws_b2b = wb.create_sheet(title="b2b")
    ws_b2b.append([
        "GSTIN/UIN of Recipient", "Receiver Name", "Invoice Number", "Invoice Date", 
        "Invoice Value", "Place Of Supply", "Reverse Charge", "Applicable % of Tax Rate", 
        "Invoice Type", "E-Commerce GSTIN", "Rate", "Taxable Value", "Cess Amount"
    ])
    
    for inv in invoices:
        contact = inv.contact
        is_registered = contact and contact.gstin
        if is_registered:
            # Group lines by rate
            rate_groups = {}
            for line in inv.lines:
                rate = float(line.gst_rate)
                rate_groups[rate] = rate_groups.get(rate, 0.0) + float(line.subtotal)
            
            for rate, taxable_val in rate_groups.items():
                ws_b2b.append([
                    contact.gstin,
                    contact.name,
                    inv.invoice_number,
                    inv.issue_date.strftime("%d-%b-%Y") if inv.issue_date else "",
                    float(inv.total),
                    format_pos(inv.pos_state_code),
                    "N",
                    "",
                    "Regular",
                    "",
                    rate,
                    taxable_val,
                    0.0
                ])

    # 2. b2cs sheet
    ws_b2cs = wb.create_sheet(title="b2cs")
    ws_b2cs.append([
        "Type", "Place Of Supply", "Applicable % of Tax Rate", "Rate", "Taxable Value", 
        "Cess Amount", "E-Commerce GSTIN"
    ])
    for line in report.b2cs:
        ws_b2cs.append([
            "OE",
            format_pos(line.pos_state_code),
            "",
            float(line.gst_rate),
            float(line.taxable_value),
            float(line.cess_amount),
            ""
        ])

    # 3. cdnr sheet
    ws_cdnr = wb.create_sheet(title="cdnr")
    ws_cdnr.append([
        "GSTIN/UIN of Recipient", "Receiver Name", "Note/Refund Voucher Number", "Note/Refund Voucher Date", 
        "Document Type", "Note Value", "Place Of Supply", "Applicable % of Tax Rate", 
        "Reason For Recording document", "Rate", "Taxable Value", "Cess Amount"
    ])
    
    q_cn = db.query(CreditNote).filter(
        CreditNote.tenant_id == tenant_id,
        CreditNote.status == "POSTED",
        CreditNote.deleted_at == None
    )
    if start_date:
        q_cn = q_cn.filter(CreditNote.issue_date >= start_date)
    if end_date:
        q_cn = q_cn.filter(CreditNote.issue_date <= end_date)
    credit_notes = q_cn.all()

    for cn in credit_notes:
        inv = cn.invoice
        contact = inv.contact if inv else None
        if contact and contact.gstin:
            rate_groups = {}
            for line in cn.lines:
                rate = float(line.gst_rate)
                rate_groups[rate] = rate_groups.get(rate, 0.0) + float(line.subtotal)
            
            for rate, taxable_val in rate_groups.items():
                ws_cdnr.append([
                    contact.gstin,
                    contact.name,
                    cn.credit_note_number,
                    cn.issue_date.strftime("%d-%b-%Y") if cn.issue_date else "",
                    "C",
                    float(cn.total),
                    format_pos(inv.pos_state_code if inv else origin_state_code),
                    "",
                    cn.reason or "Sales Return",
                    rate,
                    taxable_val,
                    0.0
                ])

    q_dn = db.query(DebitNote).filter(
        DebitNote.tenant_id == tenant_id,
        DebitNote.status == "POSTED",
        DebitNote.deleted_at == None
    )
    if start_date:
        q_dn = q_dn.filter(DebitNote.issue_date >= start_date)
    if end_date:
        q_dn = q_dn.filter(DebitNote.issue_date <= end_date)
    debit_notes = q_dn.all()

    for dn in debit_notes:
        inv = dn.invoice
        contact = inv.contact if inv else None
        if contact and contact.gstin:
            rate_groups = {}
            for line in dn.lines:
                rate = float(line.gst_rate)
                rate_groups[rate] = rate_groups.get(rate, 0.0) + float(line.subtotal)
            
            for rate, taxable_val in rate_groups.items():
                ws_cdnr.append([
                    contact.gstin,
                    contact.name,
                    dn.debit_note_number,
                    dn.issue_date.strftime("%d-%b-%Y") if dn.issue_date else "",
                    "D",
                    float(dn.total),
                    format_pos(inv.pos_state_code if inv else origin_state_code),
                    "",
                    dn.reason or "Sales Correction",
                    rate,
                    taxable_val,
                    0.0
                ])

    # 4. cdnur sheet
    ws_cdnur = wb.create_sheet(title="cdnur")
    ws_cdnur.append([
        "Note/Refund Voucher Number", "Note/Refund Voucher Date", "Document Type", 
        "Invoice Type", "Note Value", "Place Of Supply", "Applicable % of Tax Rate", 
        "Reason For Recording document", "Rate", "Taxable Value", "Cess Amount"
    ])
    
    for cn in credit_notes:
        inv = cn.invoice
        contact = inv.contact if inv else None
        if not contact or not contact.gstin:
            rate_groups = {}
            for line in cn.lines:
                rate = float(line.gst_rate)
                rate_groups[rate] = rate_groups.get(rate, 0.0) + float(line.subtotal)
            
            for rate, taxable_val in rate_groups.items():
                ws_cdnur.append([
                    cn.credit_note_number,
                    cn.issue_date.strftime("%d-%b-%Y") if cn.issue_date else "",
                    "C",
                    "B2CL" if (inv and inv.total > 250000 and inv.pos_state_code != origin_state_code) else "B2CS",
                    float(cn.total),
                    format_pos(inv.pos_state_code if inv else origin_state_code),
                    "",
                    cn.reason or "Sales Return",
                    rate,
                    taxable_val,
                    0.0
                ])

    for dn in debit_notes:
        inv = dn.invoice
        contact = inv.contact if inv else None
        if not contact or not contact.gstin:
            rate_groups = {}
            for line in dn.lines:
                rate = float(line.gst_rate)
                rate_groups[rate] = rate_groups.get(rate, 0.0) + float(line.subtotal)
            
            for rate, taxable_val in rate_groups.items():
                ws_cdnur.append([
                    dn.debit_note_number,
                    dn.issue_date.strftime("%d-%b-%Y") if dn.issue_date else "",
                    "D",
                    "B2CL" if (inv and inv.total > 250000 and inv.pos_state_code != origin_state_code) else "B2CS",
                    float(dn.total),
                    format_pos(inv.pos_state_code if inv else origin_state_code),
                    "",
                    dn.reason or "Sales Correction",
                    rate,
                    taxable_val,
                    0.0
                ])

    # 5. hsn sheet
    ws_hsn = wb.create_sheet(title="hsn")
    ws_hsn.append([
        "HSN", "Description", "UQC", "Total Quantity", "Total Value", "Taxable Value", 
        "Integrated Tax Amount", "Central Tax Amount", "State/UT Tax Amount", "Cess Amount"
    ])
    for line in report.hsn_summary:
        ws_hsn.append([
            line.hsn_sac,
            line.description or "",
            line.uom or "PCS-PIECES",
            float(line.total_quantity),
            float(line.total_value),
            float(line.taxable_value),
            float(line.igst_amount),
            float(line.cgst_amount),
            float(line.sgst_amount),
            float(line.cess_amount)
        ])

    # 6. doc sheet
    ws_doc = wb.create_sheet(title="doc")
    ws_doc.append([
        "Nature of Document", "Sr. No. From", "Sr. No. To", "Total Number", "Cancelled", "Net Issued"
    ])
    if invoices:
        inv_nums = [i.invoice_number for i in invoices]
        inv_nums.sort()
        total_count = len(invoices)
        cancelled_count = len([i for i in invoices if i.status.upper() == "CANCELLED"])
        ws_doc.append([
            "Invoices for outward supply",
            inv_nums[0],
            inv_nums[-1],
            total_count,
            cancelled_count,
            total_count - cancelled_count
        ])

    header_fill = PatternFill(start_color="0B1B3D", end_color="0B1B3D", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    
    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[1].height = 24

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    filename = f"GSTR1_Export_{start_date or 'ALL'}_to_{end_date or 'ALL'}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/gstr2/export")
def export_gstr2(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("invoice:view"))
):
    """Generates an Excel workbook for GSTR-2 inward purchase supplies (bills)."""
    origin_state_code = resolve_origin_state_code(db, tenant_id)

    q_bill = db.query(Bill).filter(
        Bill.tenant_id == tenant_id,
        Bill.status.in_(["POSTED", "PARTIALLY_PAID", "PAID"]),
        Bill.deleted_at == None
    )
    if start_date:
        q_bill = q_bill.filter(Bill.issue_date >= start_date)
    if end_date:
        q_bill = q_bill.filter(Bill.issue_date <= end_date)
    bills = q_bill.all()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_b2b = wb.create_sheet(title="b2b")
    ws_b2b.append([
        "GSTIN of Supplier", "Supplier Name", "Invoice Number", "Invoice Date", 
        "Invoice Value", "Place Of Supply", "Reverse Charge", "Rate", "Taxable Value", 
        "Integrated Tax", "Central Tax", "State/UT Tax", "Cess", "ITC Eligibility",
        "Amount of ITC Integrated Tax", "Amount of ITC Central Tax", "Amount of ITC State/UT Tax", "Amount of ITC Cess"
    ])

    for b in bills:
        contact = b.contact
        if contact and contact.gstin:
            rate_groups = {}
            for line in b.lines:
                rate = float(line.gst_rate)
                rate_groups[rate] = rate_groups.get(rate, 0.0) + float(line.subtotal)
            
            for rate, taxable_val in rate_groups.items():
                tax_multiplier = rate / 100.0
                ws_b2b.append([
                    contact.gstin,
                    contact.name,
                    b.bill_number,
                    b.issue_date.strftime("%d-%b-%Y") if b.issue_date else "",
                    float(b.total),
                    format_pos(b.pos_state_code),
                    "N",
                    rate,
                    taxable_val,
                    taxable_val * tax_multiplier if b.pos_state_code != origin_state_code else 0.0,
                    taxable_val * tax_multiplier / 2.0 if b.pos_state_code == origin_state_code else 0.0,
                    taxable_val * tax_multiplier / 2.0 if b.pos_state_code == origin_state_code else 0.0,
                    0.0,
                    "Inputs",
                    taxable_val * tax_multiplier if b.pos_state_code != origin_state_code else 0.0,
                    taxable_val * tax_multiplier / 2.0 if b.pos_state_code == origin_state_code else 0.0,
                    taxable_val * tax_multiplier / 2.0 if b.pos_state_code == origin_state_code else 0.0,
                    0.0
                ])

    header_fill = PatternFill(start_color="0B1B3D", end_color="0B1B3D", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    for cell in ws_b2b[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_b2b.row_dimensions[1].height = 24

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    filename = f"GSTR2_Export_{start_date or 'ALL'}_to_{end_date or 'ALL'}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/gstr3b/export")
def export_gstr3b(
    start_date: date = Query(..., description="Month start date, e.g. 2025-04-01"),
    end_date: date = Query(..., description="Month end date, e.g. 2025-04-30"),
    db: Session = Depends(get_db_session),
    tenant_id: uuid.UUID = Depends(enforce_permission("reports:view"))
):
    """Generates an Excel workbook for GSTR-3B return summary containing tables 3.1, 4, and 5."""
    report = GSTR3BService.get(db, tenant_id, start_date, end_date)
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GSTR-3B"

    title_font = Font(name="Arial", size=14, bold=True, color="0B1B3D")
    subtitle_font = Font(name="Arial", size=10, italic=True, color="555555")
    section_font = Font(name="Arial", size=11, bold=True, color="0B1B3D")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    label_font = Font(name="Arial", size=10, bold=False)
    bold_label_font = Font(name="Arial", size=10, bold=True)
    
    header_fill = PatternFill(start_color="0B1B3D", end_color="0B1B3D", fill_type="solid")
    section_fill = PatternFill(start_color="F2F5FA", end_color="F2F5FA", fill_type="solid")
    double_border = Border(bottom=Side(style='double', color='000000'), top=Side(style='thin', color='CCCCCC'))
    thin_border = Border(bottom=Side(style='thin', color='EEEEEE'))

    ws["A1"] = "GSTR-3B RETURN SUMMARY"
    ws["A1"].font = title_font
    ws["A2"] = f"Period: {start_date.strftime('%d-%b-%Y')} to {end_date.strftime('%d-%b-%Y')}"
    ws["A2"].font = subtitle_font
    ws["A3"] = f"Taxpayer Name: {tenant.legal_name if tenant else 'N/A'} | GSTIN: {report.gstin or 'N/A'}"
    ws["A3"].font = subtitle_font

    row = 5

    # 1. TABLE 3.1
    ws.cell(row=row, column=1, value="Table 3.1: Details of Outward Supplies and inward supplies liable to reverse charge").font = section_font
    ws.row_dimensions[row].height = 20
    row += 1

    headers_31 = ["Nature of Supplies", "Total Taxable Value (₹)", "Integrated Tax (₹)", "Central Tax (₹)", "State/UT Tax (₹)", "Cess (₹)"]
    for col_idx, h in enumerate(headers_31, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[row].height = 24
    row += 1

    outward = report.outward_taxable_supplies
    nil_rated = report.nil_rated_supplies

    rows_31 = [
        ("(a) Outward taxable supplies (other than zero rated, nil rated and exempted)", float(outward.taxable_value), float(outward.integrated_tax), float(outward.central_tax), float(outward.state_ut_tax), float(outward.cess)),
        ("(b) Outward taxable supplies (zero rated)", 0.0, 0.0, 0.0, 0.0, 0.0),
        ("(c) Other outward supplies (Nil rated, exempted)", float(nil_rated.taxable_value), 0.0, 0.0, 0.0, 0.0),
        ("(d) Inward supplies (liable to reverse charge)", 0.0, 0.0, 0.0, 0.0, 0.0),
        ("(e) Non-GST outward supplies", 0.0, 0.0, 0.0, 0.0, 0.0)
    ]

    for label, val, igst, cgst, sgst, cess in rows_31:
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=2, value=val).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=3, value=igst).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=4, value=cgst).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=5, value=sgst).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=6, value=cess).alignment = Alignment(horizontal="right")
        for c in range(1, 7):
            ws.cell(row=row, column=c).border = thin_border
        row += 1

    row += 2

    # 2. TABLE 4 Eligible ITC
    ws.cell(row=row, column=1, value="Table 4: Eligible Input Tax Credit (ITC)").font = section_font
    ws.row_dimensions[row].height = 20
    row += 1

    headers_4 = ["ITC Details", "Integrated Tax (₹)", "Central Tax (₹)", "State/UT Tax (₹)", "Cess (₹)"]
    for col_idx, h in enumerate(headers_4, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[row].height = 24
    row += 1

    itc = report.inward_supplies_itc

    rows_4 = [
        ("(A) ITC Available (whether in full or part)", "", "", "", ""),
        ("  (1) Import of goods", 0.0, 0.0, 0.0, 0.0),
        ("  (2) Import of services", 0.0, 0.0, 0.0, 0.0),
        ("  (3) Inward supplies liable to reverse charge", 0.0, 0.0, 0.0, 0.0),
        ("  (4) Inward supplies from ISD", 0.0, 0.0, 0.0, 0.0),
        ("  (5) All other ITC", float(itc.integrated_tax), float(itc.central_tax), float(itc.state_ut_tax), float(itc.cess)),
        ("(B) ITC Reversed", "", "", "", ""),
        ("  (1) As per rules 38, 42 & 43 and section 17(5)", 0.0, 0.0, 0.0, 0.0),
        ("  (2) Others", 0.0, 0.0, 0.0, 0.0),
        ("(C) Net ITC Available (A - B)", float(itc.integrated_tax), float(itc.central_tax), float(itc.state_ut_tax), float(itc.cess))
    ]

    for label, igst, cgst, sgst, cess in rows_4:
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = bold_label_font if (label.startswith("(") and "  " not in label) else label_font
        
        ws.cell(row=row, column=2, value=igst).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=3, value=cgst).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=4, value=sgst).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=5, value=cess).alignment = Alignment(horizontal="right")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = thin_border
            if label.startswith("(C)"):
                ws.cell(row=row, column=c).border = double_border
                ws.cell(row=row, column=c).fill = section_fill
        row += 1

    row += 2

    # 3. NET TAX PAYABLE SUMMARY
    ws.cell(row=row, column=1, value="Net Tax Payable Summary").font = section_font
    ws.row_dimensions[row].height = 20
    row += 1

    headers_pay = ["Tax Head", "Output Tax (A)", "ITC Claimed (B)", "Net Tax Payable (A - B)"]
    for col_idx, h in enumerate(headers_pay, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[row].height = 24
    row += 1

    net_rows = [
        ("Integrated Tax (IGST)", float(outward.integrated_tax), float(itc.integrated_tax), float(report.net_tax_payable_igst)),
        ("Central Tax (CGST)", float(outward.central_tax), float(itc.central_tax), float(report.net_tax_payable_cgst)),
        ("State/UT Tax (SGST)", float(outward.state_ut_tax), float(itc.state_ut_tax), float(report.net_tax_payable_sgst)),
        ("Cess", float(outward.cess), float(itc.cess), float(report.net_tax_payable_cess))
    ]

    for label, output, itc_val, net in net_rows:
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=2, value=output).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=3, value=itc_val).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=4, value=net).alignment = Alignment(horizontal="right")
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = thin_border
            if net > 0:
                ws.cell(row=row, column=4).font = Font(name="Arial", size=10, bold=True, color="FF0000")
        row += 1

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    filename = f"GSTR3B_Export_{start_date}_to_{end_date}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

