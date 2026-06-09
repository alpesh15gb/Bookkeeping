import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from src.schemas.report_schemas import CashBookResponse

def generate_cash_book_excel(data: CashBookResponse) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Book"

    ws.merge_cells('A1:J1')
    ws['A1'] = f"Cash Book ({data.period_start} to {data.period_end})"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E2E8F0")

    headers = ["Date", "Transaction Details", "Invoice Amount", "Tax Amount", "Amount"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num + 5)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill

    max_len = max(len(data.inflows), len(data.outflows))
    current_row = 4
    for i in range(max_len):
        if i < len(data.inflows):
            inf = data.inflows[i]
            ws.cell(row=current_row, column=1, value=str(inf.date))
            ws.cell(row=current_row, column=2, value=inf.transaction_details)
            ws.cell(row=current_row, column=3, value=float(inf.invoice_amount) if inf.invoice_amount else "")
            ws.cell(row=current_row, column=4, value=float(inf.tax_amount) if inf.tax_amount else "")
            ws.cell(row=current_row, column=5, value=float(inf.amount))
            
        if i < len(data.outflows):
            outf = data.outflows[i]
            ws.cell(row=current_row, column=6, value=str(outf.date))
            ws.cell(row=current_row, column=7, value=outf.transaction_details)
            ws.cell(row=current_row, column=8, value=float(outf.invoice_amount) if outf.invoice_amount else "")
            ws.cell(row=current_row, column=9, value=float(outf.tax_amount) if outf.tax_amount else "")
            ws.cell(row=current_row, column=10, value=float(outf.amount))
        
        current_row += 1

    ws.cell(row=current_row, column=4, value="Total").font = Font(bold=True)
    ws.cell(row=current_row, column=5, value=float(data.summary.cash_inflow)).font = Font(bold=True)
    ws.cell(row=current_row, column=9, value="Total").font = Font(bold=True)
    ws.cell(row=current_row, column=10, value=float(data.summary.cash_outflow)).font = Font(bold=True)
    
    for col in range(1, 11):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def generate_cash_book_pdf(data: CashBookResponse, company_name: str) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=10*mm, rightMargin=10*mm, topMargin=10*mm, bottomMargin=10*mm)
    
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    title_style.alignment = 1
    
    elements = []
    elements.append(Paragraph(f"<b>{company_name} - Cash Book</b>", title_style))
    elements.append(Paragraph(f"Period: {data.period_start} to {data.period_end}", styles['Normal']))
    elements.append(Spacer(1, 5*mm))
    
    table_data = []
    headers = [
        "In Date", "In Details", "In InvAmt", "In Tax", "In Amount", 
        "Out Date", "Out Details", "Out InvAmt", "Out Tax", "Out Amount"
    ]
    table_data.append(headers)
    
    max_len = max(len(data.inflows), len(data.outflows))
    for i in range(max_len):
        row = []
        if i < len(data.inflows):
            inf = data.inflows[i]
            row.extend([str(inf.date), inf.transaction_details, f"{inf.invoice_amount:.2f}" if inf.invoice_amount else "", f"{inf.tax_amount:.2f}" if inf.tax_amount else "", f"{inf.amount:.2f}"])
        else:
            row.extend(["", "", "", "", ""])
            
        if i < len(data.outflows):
            outf = data.outflows[i]
            row.extend([str(outf.date), outf.transaction_details, f"{outf.invoice_amount:.2f}" if outf.invoice_amount else "", f"{outf.tax_amount:.2f}" if outf.tax_amount else "", f"{outf.amount:.2f}"])
        else:
            row.extend(["", "", "", "", ""])
            
        table_data.append(row)
        
    table_data.append([
        "", "", "", "Total In", f"{data.summary.cash_inflow:.2f}",
        "", "", "", "Total Out", f"{data.summary.cash_outflow:.2f}"
    ])
    
    col_widths = [20*mm, 45*mm, 20*mm, 15*mm, 25*mm, 20*mm, 45*mm, 20*mm, 15*mm, 25*mm]
    t = Table(table_data, colWidths=col_widths)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#E2E8F0')),
        ('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
        ('BOX', (0,0), (-1,-1), 0.25, colors.black),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('WORDWRAP', (0,0), (-1,-1), True),
    ]))
    elements.append(t)
    
    elements.append(Spacer(1, 10*mm))
    elements.append(Paragraph(f"<b>Opening Balance:</b> {data.opening_balance:.2f}", styles['Normal']))
    elements.append(Paragraph(f"<b>Closing Balance:</b> {data.summary.closing_balance:.2f}", styles['Normal']))
    
    doc.build(elements)
    return buffer.getvalue()
